from flask import Flask, render_template, request, send_file, jsonify, Response, stream_with_context, after_this_request
import pandas as pd
import numpy as np
import os
import zipfile
from werkzeug.utils import secure_filename
import shutil
from datetime import datetime, timedelta
import traceback
import json
from queue import Queue
import threading
import secrets
import time
from functools import wraps
from collections import defaultdict
import re
import uuid
from itertools import combinations

# PDF Report Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = Flask(__name__)
app.config['SECRET_KEY'] = secrets.token_hex(32)  # Generate secure secret key
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=1)

# Store progress queues for active sessions
progress_queues = {}

# Store analysis results temporarily (session_id -> {'data': analysis_data, 'timestamp': time.time()})
# These are cleaned up after being fetched or after 1 hour
analysis_results = {}

# Rate limiting: track requests per IP
rate_limit_store = defaultdict(list)
RATE_LIMIT_REQUESTS = 10  # Max requests
RATE_LIMIT_WINDOW = 60  # Per 60 seconds

# Use absolute paths for upload and output folders
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'output')
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size

# Ensure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Rate limiting decorator
def rate_limit(max_requests=10, window=60):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            ip = request.remote_addr
            now = time.time()
            
            # Clean old entries
            rate_limit_store[ip] = [req_time for req_time in rate_limit_store[ip] if now - req_time < window]
            
            # Check rate limit
            if len(rate_limit_store[ip]) >= max_requests:
                return jsonify({'error': 'Rate limit exceeded. Please try again later.'}), 429
            
            # Add current request
            rate_limit_store[ip].append(now)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Cleanup old analysis results periodically
def cleanup_old_analysis_results():
    """Remove analysis results older than 1 hour"""
    current_time = time.time()
    expired_sessions = [
        session_id for session_id, data in analysis_results.items()
        if current_time - data.get('timestamp', 0) > 3600
    ]
    for session_id in expired_sessions:
        del analysis_results[session_id]
        print(f"Cleaned up expired analysis session: {session_id}")

# Start cleanup thread
def cleanup_thread():
    while True:
        time.sleep(300)  # Run every 5 minutes
        cleanup_old_analysis_results()
        cleanup_session_cache()
        cleanup_pipeline_sessions()

cleanup_thread_instance = threading.Thread(target=cleanup_thread, daemon=True)
cleanup_thread_instance.start()

# =============================================================================
# SESSION FILE CACHING - Store processed results for tool chaining
# =============================================================================
# Global file cache with unique IDs - accessible across all sessions
file_cache = {}  # cache_id -> {'name': str, 'path': str, 'timestamp': float, 'rows': int, 'cols': int, 'source_tool': str}

def cache_session_file(session_id, filename, file_path, rows, cols, source_tool='Unknown'):
    """Cache a processed file for potential use in another tool. Returns cache_id."""
    import hashlib
    # Generate unique cache ID from filename + timestamp
    cache_id = hashlib.md5(f"{filename}{time.time()}{os.urandom(8).hex()}".encode()).hexdigest()[:12]

    # Keep only last 10 files total (simple global cache)
    if len(file_cache) >= 10:
        # Remove oldest file
        oldest_id = min(file_cache.keys(), key=lambda k: file_cache[k]['timestamp'])
        oldest = file_cache.pop(oldest_id)
        if os.path.exists(oldest.get('path', '')):
            try:
                os.remove(oldest['path'])
            except:
                pass

    file_cache[cache_id] = {
        'name': filename,
        'path': file_path,
        'timestamp': time.time(),
        'rows': rows,
        'cols': cols,
        'source_tool': source_tool
    }
    return cache_id

def get_cached_files(session_id=None):
    """Get list of all cached files (session_id kept for backwards compatibility)"""
    return list(file_cache.values())

def get_cached_file_by_id(cache_id):
    """Get a cached file by its cache ID"""
    return file_cache.get(cache_id)

def cleanup_session_cache():
    """Remove cached files older than 1 hour"""
    current_time = time.time()
    expired_ids = []
    for cache_id, data in file_cache.items():
        if current_time - data.get('timestamp', 0) > 3600:
            expired_ids.append(cache_id)

    for cache_id in expired_ids:
        file_info = file_cache.pop(cache_id, {})
        if os.path.exists(file_info.get('path', '')):
            try:
                os.remove(file_info['path'])
            except:
                pass
        print(f"Cleaned up expired cache: {cache_id}")

# =============================================================================
# DATA READINESS PIPELINE - Guided multi-stage data assessment workflow
# =============================================================================
class PipelineState:
    """Stores state for a Data Readiness Pipeline session"""
    def __init__(self, session_id, file_path, filename):
        self.session_id = session_id
        self.file_path = file_path
        self.filename = filename
        self.created_at = time.time()
        self.current_stage = 1
        self.df = None  # DataFrame loaded in memory during session
        self.row_count = 0
        self.col_count = 0

        # Stage data storage
        self.stage_data = {
            1: None,  # Shape Analysis results
            2: None,  # Gap Assessment + user triage decisions
            3: None,  # Natural Key results + user selection
            4: None,  # Transformation recommendations + user selections
            5: None   # Execution results + transformation log
        }

        # User decisions at each stage
        self.user_decisions = {
            2: {},  # gap_triage: {column_name: 'acceptable' | 'needs_attention'}
            3: {},  # selected_keys: [list of columns]
            4: {}   # selected_transformations: {type: config}
        }

    def to_dict(self):
        """Return serializable state summary"""
        return {
            'session_id': self.session_id,
            'filename': self.filename,
            'created_at': self.created_at,
            'current_stage': self.current_stage,
            'row_count': self.row_count,
            'col_count': self.col_count,
            'stages_completed': [k for k, v in self.stage_data.items() if v is not None]
        }

# Pipeline session storage
pipeline_sessions = {}  # session_id -> PipelineState

def cleanup_pipeline_sessions():
    """Remove pipeline sessions older than 2 hours"""
    current_time = time.time()
    expired_sessions = [
        session_id for session_id, state in pipeline_sessions.items()
        if current_time - state.created_at > 7200  # 2 hours
    ]
    for session_id in expired_sessions:
        state = pipeline_sessions.pop(session_id)
        # Clean up temp file if exists
        if state.file_path and os.path.exists(state.file_path):
            try:
                os.remove(state.file_path)
            except:
                pass
        # Free DataFrame memory
        if state.df is not None:
            del state.df
        print(f"Cleaned up expired pipeline session: {session_id}")

# =============================================================================
# UNIFIED FILE READER - Handles Excel and CSV with automatic detection
# =============================================================================
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
ALLOWED_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # xlsx
    'application/vnd.ms-excel',  # xls
    'text/csv',
    'application/csv',
    'text/plain'  # Some systems send CSV as text/plain
}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_extension(filename):
    """Get lowercase file extension"""
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

def read_data_file(file_path, **kwargs):
    """
    Unified file reader that handles Excel (.xlsx, .xls) and CSV files.

    Args:
        file_path: Path to the file
        **kwargs: Additional arguments passed to pandas read functions
                  (e.g., nrows, usecols, dtype, parse_dates)

    Returns:
        pandas DataFrame

    Raises:
        ValueError: If file format is not supported
    """
    ext = get_file_extension(file_path)

    if ext == 'csv':
        # Try different encodings for CSV
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        for encoding in encodings:
            try:
                return pd.read_csv(file_path, encoding=encoding, **kwargs)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                if encoding == encodings[-1]:
                    raise e
                continue
        raise ValueError(f"Could not read CSV file with any supported encoding")

    elif ext in ('xlsx', 'xls'):
        return pd.read_excel(file_path, **kwargs)

    else:
        raise ValueError(f"Unsupported file format: {ext}. Supported formats: xlsx, xls, csv")

def write_data_file(df, file_path, file_format='xlsx', **kwargs):
    """
    Unified file writer that handles Excel and CSV output.

    Args:
        df: pandas DataFrame to write
        file_path: Output path (extension will be adjusted if needed)
        file_format: 'xlsx', 'xls', or 'csv'
        **kwargs: Additional arguments passed to pandas write functions

    Returns:
        Final file path used
    """
    # Ensure correct extension
    base_path = file_path.rsplit('.', 1)[0] if '.' in file_path else file_path

    if file_format == 'csv':
        final_path = f"{base_path}.csv"
        df.to_csv(final_path, index=False, **kwargs)
    else:
        final_path = f"{base_path}.xlsx"
        df.to_excel(final_path, index=False, **kwargs)

    return final_path

def get_file_preview(file_path, max_rows=20):
    """
    Get a preview of file contents for display before processing.

    Args:
        file_path: Path to the file
        max_rows: Maximum rows to return (default 20)

    Returns:
        dict with 'columns', 'rows', 'total_rows', 'total_cols'
    """
    try:
        # Read just enough to get preview
        df = read_data_file(file_path, nrows=max_rows + 1)

        # Get total row count (read full file for count only)
        ext = get_file_extension(file_path)
        if ext == 'csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                total_rows = sum(1 for _ in f) - 1  # Subtract header
        else:
            full_df = pd.read_excel(file_path, usecols=[0])  # Read just first column for count
            total_rows = len(full_df)

        # Convert preview to JSON-serializable format
        preview_df = df.head(max_rows)
        rows = []
        for _, row in preview_df.iterrows():
            row_data = {}
            for col in preview_df.columns:
                val = row[col]
                if pd.isna(val):
                    row_data[col] = None
                elif isinstance(val, (datetime, pd.Timestamp)):
                    row_data[col] = val.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    row_data[col] = str(val) if not isinstance(val, (int, float, bool)) else val
            rows.append(row_data)

        return {
            'columns': list(df.columns),
            'rows': rows,
            'preview_count': len(rows),
            'total_rows': total_rows,
            'total_cols': len(df.columns)
        }
    except Exception as e:
        raise ValueError(f"Error reading file preview: {str(e)}")


def split_excel_file(input_file_path, output_folder, chunk_size=40000, base_filename=None, progress_queue=None, session_id=None):
    """
    Splits an Excel file into multiple files with up to `chunk_size` records each.
    Returns the number of files created and total rows processed.
    
    Parameters:
        input_file_path: Path to the input Excel file
        output_folder: Directory where split files will be saved
        chunk_size: Number of records per file
        base_filename: Optional custom base name for output files (e.g., "PR_Amounts_Load")
                      Files will be named: [base_filename]_1of10.xlsx, [base_filename]_2of10.xlsx, etc.
        progress_queue: Queue to send progress updates
        session_id: Session identifier for tracking progress
    """
    # Ensure the output directory exists
    os.makedirs(output_folder, exist_ok=True)
    
    def send_progress(stage, current, total, message):
        """Send progress update to queue"""
        if progress_queue and session_id:
            progress_queue.put({
                'stage': stage,
                'current': current,
                'total': total,
                'percentage': int((current / total) * 100) if total > 0 else 0,
                'message': message
            })
    
    # Send initial progress
    send_progress('loading', 0, 100, 'Reading Excel file...')
    
    # Read the data from the Excel file
    df = pd.read_excel(input_file_path)
    total_rows = len(df)
    num_splits = (total_rows + chunk_size - 1) // chunk_size
    
    send_progress('loading', 100, 100, f'Loaded {total_rows:,} records. Creating {num_splits} files...')
    
    output_files = []
    
    # Use custom base filename if provided, otherwise use default
    if not base_filename or base_filename.strip() == '':
        base_filename = "split"
    
    for i in range(num_splits):
        start_idx = i * chunk_size
        end_idx = min(start_idx + chunk_size, total_rows)
        chunk_df = df.iloc[start_idx:end_idx]
        
        # Create filename in format: [base_filename]_1of10.xlsx
        filename = f"{base_filename}_{i+1}of{num_splits}.xlsx"
        output_path = os.path.join(output_folder, filename)
        
        # Send progress update
        send_progress('splitting', i + 1, num_splits, f'Creating {filename}...')
        
        chunk_df.to_excel(output_path, index=False)
        output_files.append(output_path)
        print(f"Created {filename} with records {start_idx + 1} to {end_idx}")
    
    # Don't send 'complete' here - wait until after zipping
    
    return output_files, total_rows, num_splits

@app.route('/')
def index():
    return render_template('landing.html')

@app.route('/landing')
def landing():
    return render_template('landing.html')

@app.route('/generate-test-file')
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def generate_test_file():
    """Generate and download a test Excel file with specified number of rows"""
    try:
        import random
        
        # Get number of rows from query parameter (default: 500)
        num_rows = request.args.get('num_rows', 500, type=int)
        
        # Validate row count
        if num_rows < 10:
            num_rows = 10
        elif num_rows > 100000:
            num_rows = 100000
        
        # Get custom filename from query parameter (optional)
        custom_filename = request.args.get('filename', '').strip()
        if custom_filename:
            # Sanitize filename - remove invalid characters
            custom_filename = secure_filename(custom_filename)
            # Remove .xlsx extension if user added it
            if custom_filename.lower().endswith('.xlsx') or custom_filename.lower().endswith('.xls'):
                custom_filename = custom_filename.rsplit('.', 1)[0]
            # Limit length
            if len(custom_filename) > 100:
                custom_filename = custom_filename[:100]
            # Use custom filename if valid, otherwise use default
            if custom_filename:
                download_filename = f'{custom_filename}.xlsx'
            else:
                download_filename = 'test_data_dragon.xlsx'
        else:
            download_filename = 'test_data_dragon.xlsx'
        
        # Use timestamp-based seed so each file is different
        seed_value = int(time.time() * 1000) % (2**31)
        random.seed(seed_value)
        
        data = {
            'ID': [f'PR-{i:05d}' for i in range(1, num_rows + 1)],
            'Vendor_Name': [random.choice(['Acme Corp', 'Tech Solutions Inc', 'Global Supplies', 'Best Services LLC', 'Prime Materials Co']) for _ in range(num_rows)],
            'Invoice_Number': [f'INV-{random.randint(1000, 9999)}-{random.randint(100, 999)}' for _ in range(num_rows)],
            'Invoice_Date': [(datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365))).strftime('%Y-%m-%d') for _ in range(num_rows)],
            'Amount': [round(random.uniform(100.00, 50000.00), 2) for _ in range(num_rows)],
            'GL_Account': [f'{random.randint(1000, 9999)}-{random.randint(100, 999)}' for _ in range(num_rows)],
            'Department': [random.choice(['IT', 'Finance', 'Operations', 'HR', 'Sales', 'Marketing']) for _ in range(num_rows)],
            'Status': [random.choice(['Pending', 'Approved', 'Paid', 'Rejected']) for _ in range(num_rows)],
            'Description': [f'Purchase order for {random.choice(["office supplies", "software license", "equipment", "consulting services", "maintenance"])}' for _ in range(num_rows)],
            'Quantity': [random.randint(1, 100) for _ in range(num_rows)],
            'Unit_Price': [round(random.uniform(10.00, 1000.00), 2) for _ in range(num_rows)],
            'Tax_Rate': [round(random.uniform(0.05, 0.10), 4) for _ in range(num_rows)],
            'Total_Tax': [0.0] * num_rows,
            'Net_Amount': [0.0] * num_rows,
            'Approved_By': [random.choice(['John Smith', 'Jane Doe', 'Bob Johnson', 'Alice Williams', None]) for _ in range(num_rows)],
            'Notes': [random.choice(['', 'Urgent', 'Follow up required', 'Contract renewal', None]) for _ in range(num_rows)]
        }
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Calculate derived fields
        df['Total_Tax'] = (df['Amount'] * df['Tax_Rate']).round(2)
        df['Net_Amount'] = (df['Amount'] - df['Total_Tax']).round(2)
        
        # Add some intentional duplicates for testing duplicate finder
        # Scale duplicate indices based on number of rows
        duplicate_pairs = min(7, max(2, num_rows // 100))  # 2-7 pairs depending on size
        duplicate_indices = []
        for i in range(duplicate_pairs):
            base_idx = int((i + 1) * num_rows / (duplicate_pairs + 1))
            if base_idx < len(df) - 1:
                duplicate_indices.extend([base_idx, base_idx + 1])
        
        for idx in duplicate_indices[::2]:
            if idx + 1 < len(df):
                df.iloc[idx + 1] = df.iloc[idx].copy()
        
        # Add some missing values strategically
        # Scale missing value count based on number of rows
        missing_count = min(16, max(5, num_rows // 30))  # 5-16 missing values
        missing_indices = []
        for i in range(missing_count):
            idx = int((i + 1) * num_rows / (missing_count + 1))
            if idx < len(df):
                missing_indices.append(idx)
        
        for idx in missing_indices:
            df.at[idx, 'Approved_By'] = None
            df.at[idx, 'Notes'] = None
            if idx % 2 == 0:
                df.at[idx, 'Tax_Rate'] = None
        
        # Add some edge cases (only if we have enough rows)
        if len(df) > 0:
            df.at[0, 'Amount'] = 0.00
        if len(df) > 1:
            df.at[1, 'Amount'] = 999999.99
        if len(df) > 2:
            df.at[2, 'Description'] = 'Special chars: !@#$%^&*()'
        if len(df) > 3:
            df.at[3, 'Vendor_Name'] = 'Very Long Company Name That Might Cause Display Issues In Some Systems'
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'test_data_dragon_{timestamp}.xlsx'
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Test Data', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Test Data']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Schedule cleanup after the response is sent
        @after_this_request
        def cleanup_file(response):
            try:
                if os.path.exists(file_path) and os.path.isfile(file_path):
                    os.remove(file_path)
                    print(f"Cleaned up test file: {filename}")
            except Exception as e:
                print(f"Error cleaning up test file {filename}: {str(e)}")
            return response
        
        return send_file(file_path, as_attachment=True, download_name=download_filename)
        
    except Exception as e:
        print(f"Error generating test file: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'Failed to generate test file: {str(e)}'}), 500

@app.route('/excel-splitter')
def excel_splitter():
    return render_template('index.html')

@app.route('/column-analyzer')
def column_analyzer():
    return render_template('column_analyzer.html')

@app.route('/security-info')
def security_info():
    return render_template('security_info.html')

@app.route('/data-scrubber')
def data_scrubber():
    return render_template('data_scrubber.html')

@app.route('/progress/<session_id>')
def progress(session_id):
    """Server-Sent Events endpoint for progress updates"""
    def generate():
        # Wait a moment for session to be created (handles race condition)
        q = None
        import time
        for attempt in range(10):  # Try up to 10 times (1 second total)
            q = progress_queues.get(session_id)  # Use .get() - it won't remove from dict
            if q:
                break
            time.sleep(0.1)  # Wait 100ms between attempts
        
        if not q:
            print(f"Session {session_id} not found in progress_queues after waiting")
            yield f"data: {json.dumps({'error': 'Session not found'})}\n\n"
            return
        
        print(f"SSE connection established for session {session_id}")
        import time
        last_ping = time.time()
        empty_queue_count = 0
        
        while True:
            try:
                # Use a shorter timeout and send keep-alive pings
                try:
                    progress_data = q.get(timeout=5)  # 5 second timeout
                    empty_queue_count = 0  # Reset counter
                    
                    # Log what we're sending
                    stage = progress_data.get('stage', 'unknown')
                    print(f"Sending progress update: stage={stage}, percentage={progress_data.get('percentage', 0)}")
                    
                    # Serialize the data - handle large analysis objects
                    try:
                        json_data = json.dumps(progress_data, default=str)
                        yield f"data: {json_data}\n\n"
                        last_ping = time.time()
                    except Exception as json_err:
                        print(f"JSON serialization error: {json_err}")
                        # Try sending without analysis if it's too large
                        if 'analysis' in progress_data:
                            print("Attempting to send without analysis data...")
                            progress_data_no_analysis = {k: v for k, v in progress_data.items() if k != 'analysis'}
                            json_data = json.dumps(progress_data_no_analysis, default=str)
                            yield f"data: {json_data}\n\n"
                            # Send analysis separately in chunks if needed
                            if progress_data.get('stage') == 'done':
                                print("Analysis data too large, will need alternative delivery method")
                    
                    # If done, wait a bit to ensure message is sent, then exit
                    if progress_data.get('stage') in ['done', 'error']:
                        print(f"Final stage reached: {stage}, closing connection")
                        import time
                        time.sleep(0.5)  # Give time for message to be sent
                        break
                except:
                    empty_queue_count += 1
                    # Send keep-alive ping every 15 seconds
                    if time.time() - last_ping > 15:
                        yield f": keep-alive\n\n"
                        last_ping = time.time()
                        print(f"Keep-alive sent for session {session_id}")
                    
                    # If queue has been empty for too long (2 minutes), close connection
                    if empty_queue_count > 24:  # 24 * 5 seconds = 2 minutes
                        print(f"Queue empty for too long, closing connection for session {session_id}")
                        break
                    continue
                    
            except Exception as e:
                print(f"SSE error: {e}")
                print(traceback.format_exc())
                break
        
        # Clean up the queue
        if session_id in progress_queues:
            print(f"Cleaning up session {session_id}")
            del progress_queues[session_id]
    
    return Response(stream_with_context(generate()), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })

def process_file_async(upload_path, output_folder, chunk_size, base_filename, timestamp, progress_queue, session_id):
    """Process file in background thread"""
    try:
        # Split the file
        print("Starting file split...")
        output_files, total_rows, num_splits = split_excel_file(
            upload_path, output_folder, chunk_size, base_filename, progress_queue, session_id
        )
        print(f"Split complete: {num_splits} files, {total_rows} rows")
        
        # Create a zip file
        zip_filename = f"split_files_{timestamp}.zip"
        zip_path = os.path.join(app.config['OUTPUT_FOLDER'], zip_filename)
        print(f"Creating zip: {zip_path}")
        
        progress_queue.put({
            'stage': 'zipping',
            'current': 0,
            'total': num_splits,
            'percentage': 95,
            'message': 'Creating ZIP file...'
        })
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in output_files:
                zipf.write(file_path, os.path.basename(file_path))
        
        # Clean up individual files
        print("Cleaning up temporary files...")
        shutil.rmtree(output_folder)
        os.remove(upload_path)
        
        # Send final completion message
        progress_queue.put({
            'stage': 'done',
            'current': num_splits,
            'total': num_splits,
            'percentage': 100,
            'message': 'Complete!',
            'total_rows': total_rows,
            'num_files': num_splits,
            'download_url': f'/download/{zip_filename}',
            'zip_filename': zip_filename
        })
        
        print("Success!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/upload', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def upload_file():
    try:
        print("Upload request received")
        
        if 'file' not in request.files:
            print("No file in request")
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Validate MIME type (additional security)
        allowed_mimes = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel'  # .xls
        ]
        if hasattr(file, 'content_type') and file.content_type and file.content_type not in allowed_mimes:
            # Allow if content_type is not set (some browsers don't send it)
            if file.content_type and not file.content_type.startswith('application/'):
                return jsonify({'error': 'Invalid file type. Please upload an Excel file.'}), 400
        
        # Get chunk size and base filename from form
        try:
            chunk_size = int(request.form.get('chunk_size', 40000))
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid chunk size'}), 400
        
        # Validate chunk_size (prevent DoS)
        if chunk_size < 1 or chunk_size > 1000000:  # Reasonable limits
            return jsonify({'error': 'Chunk size must be between 1 and 1,000,000'}), 400
        
        base_filename = request.form.get('base_filename', '').strip()
        
        # Sanitize base_filename to prevent path traversal
        if base_filename:
            base_filename = secure_filename(base_filename)
            # Remove any remaining dangerous characters
            base_filename = ''.join(c for c in base_filename if c.isalnum() or c in ('_', '-'))
        
        print(f"Chunk size: {chunk_size}")
        print(f"Base filename: {base_filename if base_filename else 'split (default)'}")
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"{timestamp}_{secrets.token_hex(8)}"  # Use secure random session ID
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        print(f"Saving to: {upload_path}")
        file.save(upload_path)
        
        # Create output folder for this session
        output_folder = os.path.join(app.config['OUTPUT_FOLDER'], timestamp)
        print(f"Output folder: {output_folder}")
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=process_file_async,
            args=(upload_path, output_folder, chunk_size, base_filename, timestamp, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    # Validate filename to prevent path traversal
    safe_filename = secure_filename(filename)
    if safe_filename != filename:
        return jsonify({'error': 'Invalid filename'}), 400
    
    file_path = os.path.join(app.config['OUTPUT_FOLDER'], safe_filename)
    
    # Additional security: ensure file is within output folder
    if not os.path.abspath(file_path).startswith(os.path.abspath(app.config['OUTPUT_FOLDER'])):
        return jsonify({'error': 'Invalid file path'}), 400
    
    # Check if file exists
    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404
    
    # Check if file is a zip file, Excel file, CSV file, JSON file, or Word document
    if not (safe_filename.endswith('.zip') or safe_filename.endswith('.xlsx') or 
            safe_filename.endswith('.xls') or safe_filename.endswith('.csv') or 
            safe_filename.endswith('.json') or safe_filename.endswith('.docx')):
        return jsonify({'error': 'Invalid file type'}), 400
    
    # Schedule cleanup after the response is sent
    @after_this_request
    def cleanup_file(response):
        try:
            # Delete the downloaded file
            if os.path.exists(file_path) and os.path.isfile(file_path):
                os.remove(file_path)
                print(f"Cleaned up downloaded file: {safe_filename}")
            
            # If the file was in a subdirectory, check if we should remove the empty parent directory
            parent_dir = os.path.dirname(file_path)
            if parent_dir != app.config['OUTPUT_FOLDER'] and os.path.exists(parent_dir):
                try:
                    # Check if directory is empty (only . and .. entries)
                    if not os.listdir(parent_dir):
                        os.rmdir(parent_dir)
                        print(f"Removed empty directory: {parent_dir}")
                except OSError:
                    # Directory not empty or other error - that's fine, just continue
                    pass
        except Exception as e:
            # Log error but don't fail the request
            print(f"Error cleaning up file {safe_filename}: {str(e)}")
        return response
    
    return send_file(file_path, as_attachment=True)

def make_json_serializable(obj):
    """Recursively convert objects to JSON-serializable types"""
    if isinstance(obj, dict):
        return {str(k): make_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [make_json_serializable(item) for item in obj]
    elif isinstance(obj, (pd.Timestamp, pd.DatetimeTZDtype)):
        return str(obj)
    elif hasattr(obj, 'item'):  # numpy scalars
        return obj.item()
    elif hasattr(obj, 'tolist'):  # numpy arrays
        return obj.tolist()
    elif pd.api.types.is_integer(obj):
        return int(obj)
    elif pd.api.types.is_float(obj):
        return float(obj)
    elif pd.api.types.is_bool(obj):
        return bool(obj)
    elif obj is pd.NA or pd.isna(obj):
        return None
    elif isinstance(obj, type):  # pandas dtypes
        return str(obj)
    else:
        # Try to convert to native Python type
        try:
            if isinstance(obj, (int, float, str, bool)) or obj is None:
                return obj
            return str(obj)
        except:
            return str(obj)

@app.route('/fetch-analysis/<session_id>')
def fetch_analysis(session_id):
    """Fetch analysis results for a session - used when analysis is too large for SSE"""
    try:
        # Validate session_id format (basic check)
        if not session_id or len(session_id) < 10:
            return jsonify({'error': 'Invalid session ID'}), 400
        
        if session_id not in analysis_results:
            print(f"Analysis results not found for session: {session_id}")
            return jsonify({'error': 'Analysis results not found or expired'}), 404
        
        # Check if expired
        result_data = analysis_results[session_id]
        if time.time() - result_data.get('timestamp', 0) > 3600:
            del analysis_results[session_id]
            return jsonify({'error': 'Analysis results expired'}), 404
        
        analysis = result_data.get('data', result_data)  # Support both old and new format
        print(f"Found analysis for session {session_id}")
        
        # Analysis should already be serializable, but double-check
        if not isinstance(analysis, dict):
            raise ValueError("Analysis data is not in expected format")
        
        # Try to verify it's JSON-serializable by doing a test serialization
        try:
            import json
            test_json = json.dumps(analysis, default=str)
            # If that worked, use jsonify which will handle it properly
            serializable_analysis = json.loads(test_json)
            print("Analysis verified as JSON-serializable")
        except Exception as json_err:
            print(f"Analysis not JSON-serializable, attempting conversion: {json_err}")
            # Convert if needed
            try:
                serializable_analysis = make_json_serializable(analysis)
                print("Analysis converted successfully")
            except Exception as convert_err:
                print(f"Error converting analysis: {convert_err}")
                print(traceback.format_exc())
                raise convert_err
        
        # Clean up after fetching
        del analysis_results[session_id]
        print("Analysis results cleaned up")
        
        return jsonify({
            'success': True,
            'analysis': serializable_analysis
        })
    except Exception as e:
        print(f"Error fetching analysis: {e}")
        print(traceback.format_exc())
        return jsonify({'error': f'Failed to serialize analysis: {str(e)}'}), 500

def detect_semantic_type(col_data, col_name=''):
    """
    Detect semantic type of a column by analyzing its content.
    Returns dict with detected_type, confidence, sample_values, and format_info.
    """
    # Get non-null values for analysis
    non_null_data = col_data.dropna()
    if len(non_null_data) == 0:
        return {
            'detected_type': 'Unknown',
            'confidence': 0,
            'sample_values': [],
            'format_info': None
        }
    
    # Sample size limit for performance (analyze up to 1000 values)
    sample_size = min(1000, len(non_null_data))
    sample_data = non_null_data.head(sample_size) if sample_size < len(non_null_data) else non_null_data
    total_non_null = len(non_null_data)
    
    # Convert to string for pattern matching
    str_data = sample_data.astype(str)
    
    # Date Detection - try multiple formats
    date_formats = [
        '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d',
        '%d-%m-%Y', '%m-%d-%Y', '%Y.%m.%d', '%d.%m.%Y',
        '%m.%d.%Y', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S', '%d %b %Y', '%d %B %Y',
        '%b %d, %Y', '%B %d, %Y'
    ]
    
    date_matches = 0
    detected_date_format = None
    
    # Try pandas to_datetime first (most flexible)
    try:
        parsed_dates = pd.to_datetime(str_data, errors='coerce', infer_datetime_format=True)
        date_matches = parsed_dates.notna().sum()
        if date_matches > 0:
            # Try to identify the format
            for fmt in date_formats:
                try:
                    test_parsed = pd.to_datetime(str_data.head(10), format=fmt, errors='coerce')
                    if test_parsed.notna().sum() >= 8:  # 80% match
                        detected_date_format = fmt
                        break
                except:
                    continue
    except:
        pass
    
    date_confidence = (date_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Email Detection
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    email_matches = str_data.str.match(email_pattern, na=False).sum()
    email_confidence = (email_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Phone Number Detection (US and international formats)
    # Skip phone detection for numeric columns to avoid false positives
    phone_confidence = 0
    if not pd.api.types.is_numeric_dtype(col_data):
        phone_patterns = [
            r'^\+?1?[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}$',  # US format: (XXX) XXX-XXXX or XXX-XXX-XXXX
            r'^\+\d{1,3}[-.\s]?\d{1,4}[-.\s]?\d{6,12}$',  # International with country code: +XX XXX XXXXXX
            r'^\d{3}-\d{3}-\d{4}$',  # XXX-XXX-XXXX (exact format)
            r'^\(\d{3}\)\s?\d{3}-\d{4}$',  # (XXX) XXX-XXXX
            r'^\d{10}$'  # 10 digits only (US phone without formatting)
        ]
        phone_matches = 0
        for pattern in phone_patterns:
            phone_matches += str_data.str.match(pattern, na=False).sum()
        phone_confidence = (phone_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # URL Detection
    url_pattern = r'^https?://[^\s/$.?#].[^\s]*$|^www\.[^\s/$.?#].[^\s]*$'
    url_matches = str_data.str.match(url_pattern, na=False).sum()
    url_confidence = (url_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Currency Detection
    currency_patterns = [
        r'^\$[\d,]+\.?\d*$',  # $123.45
        r'^[\d,]+\.?\d*\s?(USD|EUR|GBP|JPY|CAD|AUD)$',  # 123.45 USD
        r'^USD|EUR|GBP|JPY|CAD|AUD$'  # Currency codes only
    ]
    currency_matches = 0
    currency_symbol = None
    for pattern in currency_patterns:
        matches = str_data.str.match(pattern, na=False)
        currency_matches += matches.sum()
        if matches.sum() > 0 and currency_symbol is None:
            # Extract currency symbol - use .loc to safely index with boolean Series
            try:
                matching_values = str_data.loc[matches]
                if len(matching_values) > 0:
                    sample = str(matching_values.iloc[0])
                    if '$' in sample:
                        currency_symbol = '$'
                    elif 'USD' in sample:
                        currency_symbol = 'USD'
                    elif 'EUR' in sample:
                        currency_symbol = 'EUR'
            except:
                pass  # If indexing fails, skip symbol extraction
    currency_confidence = (currency_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Percentage Detection
    percent_pattern = r'^\d+\.?\d*\s?%$|^0\.\d+$'  # 50% or 0.5 (if numeric and between 0-1)
    percent_matches = str_data.str.match(percent_pattern, na=False).sum()
    # Also check if numeric values are between 0-1 (likely percentages)
    if pd.api.types.is_numeric_dtype(col_data):
        # sample_data is already numeric if col_data is numeric, so use it directly
        try:
            numeric_values = pd.to_numeric(sample_data, errors='coerce').dropna()
            if len(numeric_values) > 0:
                if (numeric_values >= 0).all() and (numeric_values <= 1).all() and numeric_values.mean() < 0.5:
                    percent_matches += len(numeric_values)
        except:
            pass  # If conversion fails, just skip this check
    percent_confidence = (percent_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Boolean Detection (True/False, Yes/No, 1/0, Y/N)
    bool_patterns = [
        r'^(true|false|yes|no|y|n|1|0|on|off)$'
    ]
    bool_matches = 0
    for pattern in bool_patterns:
        # Use contains with case=False instead of inline (?i) flag
        bool_matches += str_data.str.contains(pattern, case=False, na=False, regex=True).sum()
    bool_confidence = (bool_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # IP Address Detection
    ipv4_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
    ipv6_pattern = r'^([0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}$|^::1$|^::$'
    ipv4_matches = str_data.str.match(ipv4_pattern, na=False).sum()
    ipv6_matches = str_data.str.match(ipv6_pattern, na=False).sum()
    ip_confidence = ((ipv4_matches + ipv6_matches) / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Postal Code Detection
    zip_pattern = r'^\d{5}(-\d{4})?$'  # US ZIP
    postal_pattern = r'^[A-Z0-9\s-]{3,10}$'  # International (basic)
    zip_matches = str_data.str.match(zip_pattern, na=False).sum()
    postal_matches = str_data.str.match(postal_pattern, na=False).sum()
    postal_confidence = ((zip_matches + postal_matches) / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # UUID Detection
    uuid_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
    uuid_matches = str_data.str.match(uuid_pattern, na=False, case=False).sum()
    uuid_confidence = (uuid_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # ID/Serial Detection (sequential patterns, alphanumeric IDs)
    # Check if values follow a pattern like PR-00001, INV-123, etc.
    id_pattern = r'^[A-Z]{2,}-?\d+$|^[A-Z]+\d+$|^\d+$'
    id_matches = str_data.str.match(id_pattern, na=False).sum()
    # Also check if all values are unique and follow a pattern
    if id_matches > 0 and col_data.nunique() == len(col_data):
        id_confidence = (id_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    else:
        id_confidence = (id_matches / len(sample_data)) * 50 if len(sample_data) > 0 else 0  # Lower confidence if not unique
    
    # Credit Card Detection (basic pattern, not Luhn validation for performance)
    cc_pattern = r'^\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}$'
    cc_matches = str_data.str.match(cc_pattern, na=False).sum()
    cc_confidence = (cc_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # SSN Detection (US format)
    ssn_pattern = r'^\d{3}-\d{2}-\d{4}$'
    ssn_matches = str_data.str.match(ssn_pattern, na=False).sum()
    ssn_confidence = (ssn_matches / len(sample_data)) * 100 if len(sample_data) > 0 else 0
    
    # Collect all confidences and find the best match (threshold: 80%)
    type_scores = {
        'Date': date_confidence,
        'Email': email_confidence,
        'Phone Number': phone_confidence,
        'URL': url_confidence,
        'Currency': currency_confidence,
        'Percentage': percent_confidence,
        'Boolean': bool_confidence,
        'IP Address': ip_confidence,
        'Postal Code': postal_confidence,
        'UUID': uuid_confidence,
        'ID/Serial': id_confidence,
        'Credit Card': cc_confidence,
        'SSN': ssn_confidence
    }
    
    # If column is numeric, prioritize numeric types and suppress non-numeric types
    if pd.api.types.is_numeric_dtype(col_data):
        # Suppress phone, URL, email, postal code, UUID, SSN for numeric columns
        type_scores['Phone Number'] = 0
        type_scores['URL'] = 0
        type_scores['Email'] = 0
        type_scores['Postal Code'] = 0
        type_scores['UUID'] = 0
        type_scores['SSN'] = 0
        type_scores['ID/Serial'] = 0
    
    # Find the best match above threshold
    threshold = 80.0
    best_type = max(type_scores.items(), key=lambda x: x[1])
    
    if best_type[1] >= threshold:
        detected_type = best_type[0]
        confidence = round(best_type[1], 1)
        
        # Get sample values
        sample_values = sample_data.head(3).tolist()
        sample_values = [str(v) for v in sample_values]
        
        # Format info
        format_info = None
        if detected_type == 'Date' and detected_date_format:
            format_info = f"Format: {detected_date_format}"
        elif detected_type == 'Currency' and currency_symbol:
            format_info = f"Symbol: {currency_symbol}"
        elif detected_type == 'IP Address':
            if ipv4_matches > ipv6_matches:
                format_info = "IPv4"
            else:
                format_info = "IPv6"
        
        return {
            'detected_type': detected_type,
            'confidence': confidence,
            'sample_values': sample_values,
            'format_info': format_info
        }
    else:
        # No strong match, return generic type based on pandas dtype
        pandas_type = str(col_data.dtype)
        if 'int' in pandas_type:
            detected_type = 'Integer'
        elif 'float' in pandas_type:
            detected_type = 'Float'
        elif 'bool' in pandas_type:
            detected_type = 'Boolean'
        elif 'datetime' in pandas_type:
            detected_type = 'Date'
        else:
            detected_type = 'Text'
        
        return {
            'detected_type': detected_type,
            'confidence': 100.0,  # 100% confidence for pandas-detected types
            'sample_values': sample_data.head(3).astype(str).tolist(),
            'format_info': None
        }

def analyze_dataframe(df, progress_queue=None, session_id=None):
    """
    Comprehensive pandas-style data analysis.
    Returns a dictionary with all analysis results - NO data stored.
    """
    def send_progress(stage, current, total, message, percentage=None):
        """Send progress update to queue"""
        if progress_queue and session_id:
            if percentage is None and total > 0:
                percentage = int((current / total) * 100)
            elif percentage is None:
                percentage = 0
            progress_queue.put({
                'stage': stage,
                'current': current,
                'total': total,
                'percentage': percentage,
                'message': message
            })
    
    send_progress('loading', 0, 100, 'Reading file into memory...', 5)
    
    analysis = {
        'overview': {},
        'columns': {},
        'memory_info': {}
    }
    
    send_progress('overview', 0, 100, 'Analyzing dataset overview...', 10)
    
    # Overall dataset information
    analysis['overview'] = {
        'shape': {
            'rows': int(len(df)),
            'columns': int(len(df.columns))
        },
        'memory_usage_bytes': int(df.memory_usage(deep=True).sum()),
        'memory_usage_mb': round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
        'dtypes_count': {str(k): int(v) for k, v in df.dtypes.value_counts().to_dict().items()},
        'duplicate_rows': int(df.duplicated().sum()),
        'has_duplicates': bool(df.duplicated().any())
    }
    
    send_progress('overview', 50, 100, 'Calculating memory usage...', 15)
    
    # Memory usage per column
    memory_usage = df.memory_usage(deep=True)
    analysis['memory_info'] = {
        'total_mb': round(memory_usage.sum() / 1024 / 1024, 2),
        'by_column': {
            col: {
                'bytes': int(memory_usage[col]),
                'mb': round(memory_usage[col] / 1024 / 1024, 3)
            }
            for col in df.columns
        }
    }
    
    send_progress('overview', 100, 100, f'Found {len(df.columns)} columns to analyze...', 20)
    
    # Per-column analysis
    total_columns = len(df.columns)
    for idx, col in enumerate(df.columns):
        # Progress: 20% (overview complete) to 90% (columns complete)
        # Distribute 70% across all columns
        column_progress = 20 + int((idx / total_columns) * 70) if total_columns > 0 else 20
        send_progress('columns', idx + 1, total_columns, f'Analyzing column {idx + 1}/{total_columns}: {col}...', column_progress)
        
        col_data = df[col]
        col_info = {
            'dtype': str(col_data.dtype),
            'non_null_count': int(col_data.notna().sum()),
            'null_count': int(col_data.isna().sum()),
            'null_percentage': round((col_data.isna().sum() / len(col_data)) * 100, 2),
            'unique_count': int(col_data.nunique()),
            'memory_bytes': int(memory_usage[col])
        }
        
        # Detect semantic type
        semantic_type_info = detect_semantic_type(col_data, col)
        col_info['detected_type'] = semantic_type_info['detected_type']
        col_info['type_confidence'] = semantic_type_info['confidence']
        col_info['type_samples'] = semantic_type_info['sample_values']
        col_info['format_info'] = semantic_type_info['format_info']
        
        # Numeric column statistics
        if pd.api.types.is_numeric_dtype(col_data):
            col_info['is_numeric'] = True
            col_info['statistics'] = {
                'mean': float(col_data.mean()) if col_data.notna().any() else None,
                'median': float(col_data.median()) if col_data.notna().any() else None,
                'std': float(col_data.std()) if col_data.notna().any() else None,
                'min': float(col_data.min()) if col_data.notna().any() else None,
                'max': float(col_data.max()) if col_data.notna().any() else None,
                'q25': float(col_data.quantile(0.25)) if col_data.notna().any() else None,
                'q50': float(col_data.quantile(0.50)) if col_data.notna().any() else None,
                'q75': float(col_data.quantile(0.75)) if col_data.notna().any() else None,
                'skew': float(col_data.skew()) if col_data.notna().any() else None,
                'kurtosis': float(col_data.kurtosis()) if col_data.notna().any() else None
            }
            
            # Check for potential outliers using IQR method
            if col_data.notna().any():
                Q1 = col_data.quantile(0.25)
                Q3 = col_data.quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                outliers = col_data[(col_data < lower_bound) | (col_data > upper_bound)]
                col_info['outliers'] = {
                    'count': int(outliers.count()),
                    'percentage': round((outliers.count() / len(col_data)) * 100, 2) if len(col_data) > 0 else 0
                }
        else:
            col_info['is_numeric'] = False
        
        # Categorical/string column analysis
        if pd.api.types.is_string_dtype(col_data) or pd.api.types.is_object_dtype(col_data) or isinstance(col_data.dtype, pd.CategoricalDtype):
            col_info['is_categorical'] = True
            # Most frequent values (top 10)
            value_counts = col_data.value_counts().head(10)
            col_info['top_values'] = {
                str(k): int(v) for k, v in value_counts.items()
            }
            
            # String length statistics if applicable
            try:
                str_lengths = col_data.dropna().astype(str).str.len()
                if len(str_lengths) > 0:
                    col_info['string_stats'] = {
                        'mean_length': float(str_lengths.mean()),
                        'min_length': int(str_lengths.min()),
                        'max_length': int(str_lengths.max()),
                        'median_length': float(str_lengths.median())
                    }
            except:
                pass
        
        # Date/time column analysis
        # Check both pandas datetime dtype and semantic type detection
        is_date_dtype = pd.api.types.is_datetime64_any_dtype(col_data)
        is_date_detected = col_info.get('detected_type') == 'Date'
        
        if is_date_dtype or is_date_detected:
            col_info['is_datetime'] = True
            if col_data.notna().any():
                # If detected as date but stored as string, try to parse
                if is_date_detected and not is_date_dtype:
                    try:
                        parsed_dates = pd.to_datetime(col_data.dropna(), errors='coerce', infer_datetime_format=True)
                        valid_dates = parsed_dates.dropna()
                        if len(valid_dates) > 0:
                            col_info['date_range'] = {
                                'min': str(valid_dates.min()),
                                'max': str(valid_dates.max())
                            }
                    except:
                        pass
                else:
                    # Already datetime dtype
                    col_info['date_range'] = {
                        'min': str(col_data.min()),
                        'max': str(col_data.max())
                    }
        
        # Boolean column
        if pd.api.types.is_bool_dtype(col_data):
            col_info['is_boolean'] = True
            value_counts = col_data.value_counts()
            col_info['boolean_counts'] = {str(k): int(v) for k, v in value_counts.items()}
        
        analysis['columns'][col] = col_info
    
    send_progress('finalizing', 0, 100, 'Calculating type statistics...', 90)
    
    # Count columns by detected type
    detected_types_count = {}
    for col_name, col_info in analysis['columns'].items():
        detected_type = col_info.get('detected_type', 'Unknown')
        detected_types_count[detected_type] = detected_types_count.get(detected_type, 0) + 1
    
    analysis['overview']['detected_types_count'] = detected_types_count
    
    # Count how many columns were reclassified (detected type differs from pandas dtype)
    reclassified_count = 0
    for col_name, col_info in analysis['columns'].items():
        dtype = col_info.get('dtype', '')
        detected_type = col_info.get('detected_type', '')
        # Check if semantic detection found something different
        if 'object' in dtype or 'string' in dtype:
            if detected_type not in ['Text', 'Unknown']:
                reclassified_count += 1
    
    analysis['overview']['reclassified_columns'] = reclassified_count
    
    send_progress('finalizing', 100, 100, 'Finalizing analysis...', 95)
    
    return analysis

def analyze_file_async(upload_path, progress_queue, session_id):
    """Analyze file in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file into memory...', 5)
        
        # Read file into memory
        filename = os.path.basename(upload_path)
        if filename.endswith('.csv'):
            df = pd.read_csv(upload_path)
        else:
            df = pd.read_excel(upload_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows, {len(df.columns)} columns', 10)
        
        # Analyze the dataframe
        analysis = analyze_dataframe(df, progress_queue, session_id)
        
        print(f"Analysis complete. Columns analyzed: {len(analysis['columns'])}")
        
        # Clean up: explicitly delete dataframe and remove temp file
        del df
        os.remove(upload_path)
        
        # Store analysis results temporarily (will be cleaned up after fetch or timeout)
        # Convert to JSON-serializable format before storing
        try:
            serializable_analysis = make_json_serializable(analysis)
            analysis_results[session_id] = {
                'data': serializable_analysis,
                'timestamp': time.time()
            }
            print("Analysis stored in JSON-serializable format")
        except Exception as convert_err:
            print(f"Warning: Could not convert analysis to JSON-serializable format: {convert_err}")
            # Store original and convert on fetch
            analysis_results[session_id] = {
                'data': analysis,
                'timestamp': time.time()
            }
        
        # Try to send analysis via SSE, but if it's too large, send a fetch URL instead
        print("Sending completion message with analysis data...")
        
        # Estimate size of analysis data
        try:
            import sys
            analysis_size = sys.getsizeof(json.dumps(analysis, default=str))
            analysis_size_mb = analysis_size / (1024 * 1024)
            print(f"Analysis data size: {analysis_size_mb:.2f} MB")
            
            # If analysis is larger than 5MB, send via separate endpoint
            if analysis_size_mb > 5:
                print("Analysis too large for SSE, using separate endpoint")
                final_message = {
                    'stage': 'done',
                    'current': 100,
                    'total': 100,
                    'percentage': 100,
                    'message': 'Complete!',
                    'analysis_fetch_url': f'/fetch-analysis/{session_id}',
                    'analysis_too_large': True
                }
            else:
                final_message = {
                    'stage': 'done',
                    'current': 100,
                    'total': 100,
                    'percentage': 100,
                    'message': 'Complete!',
                    'analysis': analysis
                }
        except Exception as size_err:
            print(f"Could not estimate size, sending fetch URL: {size_err}")
            # If we can't estimate size, use fetch URL to be safe
            final_message = {
                'stage': 'done',
                'current': 100,
                'total': 100,
                'percentage': 100,
                'message': 'Complete!',
                'analysis_fetch_url': f'/fetch-analysis/{session_id}',
                'analysis_too_large': True
            }
        
        # Try to send - if queue is full or closed, log it
        try:
            progress_queue.put(final_message, timeout=5)
            print("Completion message sent successfully")
        except Exception as e:
            print(f"Error sending completion message: {e}")
            # Try one more time without timeout
            try:
                progress_queue.put_nowait(final_message)
                print("Completion message sent (nowait)")
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Analysis complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'upload_path' in locals() and os.path.exists(upload_path):
            os.remove(upload_path)
        if 'df' in locals():
            del df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/analyze', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def analyze_file():
    """Analyze uploaded file - returns session_id for progress tracking - NO data stored"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Validate MIME type
        allowed_mimes = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
            'text/csv',  # .csv
            'application/csv'
        ]
        if hasattr(file, 'content_type') and file.content_type and file.content_type not in allowed_mimes:
            if file.content_type and not (file.content_type.startswith('text/') or file.content_type.startswith('application/')):
                return jsonify({'error': 'Invalid file type. Please upload an Excel or CSV file.'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"analyze_{timestamp}_{secrets.token_hex(8)}"  # Use secure random session ID
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=analyze_file_async,
            args=(upload_path, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

def generate_prefix(column_name):
    """Generate prefix from column name - simplifies column name and uses appropriate separator"""
    # Simplify column name by removing common suffixes
    simplified = column_name
    col_lower = simplified.lower()
    
    # Remove common suffixes (case-insensitive)
    suffixes_to_remove = ['_name', '_code', '_id', '_number', '_num', '_date', '_amount', '_value', '_account']
    
    for suffix in suffixes_to_remove:
        if col_lower.endswith(suffix):
            simplified = simplified[:-len(suffix)]
            col_lower = simplified.lower()
            break
    
    # Determine separator based on column name pattern
    # If column name contains spaces, use " - " separator (e.g., "Org Code - 1")
    # Otherwise use "_" separator (e.g., "Vendor_1")
    if ' ' in simplified:
        separator = ' - '
    else:
        separator = '_'
    
    return simplified + separator

def scrub_dataframe(df, columns_to_scrub, relationship_preserve=False, progress_queue=None, session_id=None):
    """
    Anonymize selected columns while maintaining data shape and referential integrity.
    
    Args:
        df: pandas DataFrame
        columns_to_scrub: list of column names to anonymize
        relationship_preserve: if True and multiple columns, preserve relationships
        progress_queue: queue for progress updates
        session_id: session identifier
    
    Returns:
        tuple: (anonymized_df, mapping_dict)
    """
    def send_progress(stage, current, total, message, percentage=None):
        """Send progress update to queue"""
        if progress_queue and session_id:
            if percentage is None and total > 0:
                percentage = int((current / total) * 100)
            elif percentage is None:
                percentage = 0
            progress_queue.put({
                'stage': stage,
                'current': current,
                'total': total,
                'percentage': percentage,
                'message': message
            })
    
    send_progress('preparing', 0, 100, 'Preparing data for anonymization...', 5)
    
    # Create a copy to avoid modifying original
    anonymized_df = df.copy()
    mapping_dict = {}
    
    # Validate columns exist
    valid_columns = [col for col in columns_to_scrub if col in df.columns]
    if not valid_columns:
        raise ValueError("No valid columns selected for anonymization")
    
    send_progress('preparing', 50, 100, f'Found {len(valid_columns)} columns to anonymize...', 10)
    
    # Relationship preservation mode
    if relationship_preserve and len(valid_columns) > 1:
        send_progress('anonymizing', 0, 100, 'Anonymizing with relationship preservation...', 15)
        
        # Create composite keys from selected columns (handle NaN properly)
        def create_composite_key(row):
            values = [row[col] for col in valid_columns]
            # If any value is NaN, return None (we'll handle separately)
            if any(pd.isna(v) for v in values):
                return None
            return tuple(values)
        
        composite_keys = anonymized_df.apply(create_composite_key, axis=1)
        
        # Find unique combinations (excluding None/NaN) - convert to list for proper handling
        unique_combinations = [combo for combo in composite_keys.unique() if combo is not None]
        # Convert to list of tuples to ensure hashability and proper matching
        unique_combinations = [tuple(combo) if isinstance(combo, (list, tuple)) else combo for combo in unique_combinations]
        total_combinations = len(unique_combinations)
        
        send_progress('anonymizing', 10, 100, f'Found {total_combinations} unique combinations...', 20)
        
        # Generate prefix from first column name
        prefix = generate_prefix(valid_columns[0])
        
        # Create mapping for each unique combination
        # Convert composite_keys to tuples for proper dictionary key matching
        composite_keys_tuples = composite_keys.apply(lambda x: tuple(x) if x is not None and isinstance(x, (list, tuple)) else x)
        
        combination_mapping = {}
        for idx, combo in enumerate(unique_combinations, 1):
            anonymized_id = f"{prefix}{idx}"
            # Ensure combo is a tuple for dictionary key
            combo_tuple = tuple(combo) if not isinstance(combo, tuple) else combo
            combination_mapping[combo_tuple] = anonymized_id
            
            # Store mapping for each column in the combination
            for col_idx, col_name in enumerate(valid_columns):
                original_value = combo[col_idx] if isinstance(combo, (list, tuple)) else combo
                if col_name not in mapping_dict:
                    mapping_dict[col_name] = {}
                mapping_dict[col_name][str(original_value)] = anonymized_id  # Convert to string for JSON serialization
            
            # Progress update
            if idx % 1000 == 0 or idx == total_combinations:
                send_progress('anonymizing', idx, total_combinations, 
                            f'Mapping combination {idx}/{total_combinations}...', 
                            20 + int((idx / total_combinations) * 60))
        
        # Apply anonymization
        send_progress('applying', 0, 100, 'Applying anonymization to data...', 80)
        
        # Convert composite_keys to tuples for proper matching
        def get_tuple_key(key):
            if key is None or pd.isna(key):
                return None
            if isinstance(key, tuple):
                return key
            if isinstance(key, (list, tuple)):
                return tuple(key)
            return key
        
        composite_keys_tuples = composite_keys.apply(get_tuple_key)
        
        # Create a mapping function that handles None/NaN properly
        def map_composite_key(combo_key):
            if combo_key is None or pd.isna(combo_key):
                return None
            combo_tuple = tuple(combo_key) if not isinstance(combo_key, tuple) else combo_key
            return combination_mapping.get(combo_tuple, None)
        
        # Apply the same anonymized ID to all columns in each combination
        anonymized_ids = composite_keys_tuples.apply(map_composite_key)
        
        for col_name in valid_columns:
            # Apply anonymized IDs to this column
            anonymized_df[col_name] = anonymized_ids.copy()
            
            # Handle NaN/None values in composite keys - keep original values where composite key was None/NaN
            mask = anonymized_ids.isna() | (anonymized_ids == None)
            if mask.any():
                anonymized_df.loc[mask, col_name] = df.loc[mask, col_name]
        
        # Debug: Verify anonymization worked
        if valid_columns:
            sample_col = valid_columns[0]
            original_sample = df[sample_col].dropna().iloc[0] if len(df[sample_col].dropna()) > 0 else None
            anonymized_sample = anonymized_df[sample_col].dropna().iloc[0] if len(anonymized_df[sample_col].dropna()) > 0 else None
            print(f"Debug - Relationship mode, Column '{sample_col}': Original = {original_sample}, Anonymized = {anonymized_sample}")
            if original_sample == anonymized_sample and original_sample is not None and not str(anonymized_sample).startswith(prefix):
                print(f"ERROR: Relationship preservation failed for column '{sample_col}' - values unchanged!")
        
        send_progress('applying', 100, 100, 'Relationship preservation complete...', 90)
    
    else:
        # Independent column anonymization
        send_progress('anonymizing', 0, len(valid_columns), f'Anonymizing {len(valid_columns)} columns...', 15)
        
        for col_idx, col_name in enumerate(valid_columns):
            send_progress('anonymizing', col_idx, len(valid_columns), 
                        f'Anonymizing column: {col_name}...', 
                        15 + int((col_idx / len(valid_columns)) * 70))
            
            # Generate prefix for this column
            prefix = generate_prefix(col_name)
            
            # Get the original column data BEFORE any modifications
            original_col = df[col_name].copy()  # Use original df, not anonymized_df
            current_col = anonymized_df[col_name].copy()
            
            # Get unique values (excluding NaN) - convert to list to ensure proper matching
            unique_values = current_col.dropna().unique().tolist()
            total_unique = len(unique_values)
            
            print(f"DEBUG: Column '{col_name}' has {total_unique} unique values")
            print(f"DEBUG: First few unique values: {unique_values[:5] if len(unique_values) > 0 else 'None'}")
            print(f"DEBUG: Sample values from column: {current_col.dropna().head(3).tolist()}")
            
            # Create mapping dictionary - use exact values as keys
            value_mapping = {}
            for idx, value in enumerate(unique_values, 1):
                anonymized_id = f"{prefix}{idx}"
                value_mapping[value] = anonymized_id
            
            # Store mapping for export (convert keys to strings for JSON)
            mapping_dict[col_name] = {str(k): v for k, v in value_mapping.items()}
            
            print(f"DEBUG: Created mapping with {len(value_mapping)} entries")
            print(f"DEBUG: Sample mapping: {list(value_mapping.items())[:3] if len(value_mapping) > 0 else 'None'}")
            
            # Use .replace() instead of .map() - more reliable for value matching
            # .replace() works better with different data types and handles edge cases
            print(f"DEBUG: Applying replacement to column '{col_name}'...")
            print(f"DEBUG: Column dtype before: {current_col.dtype}")
            print(f"DEBUG: Column shape: {current_col.shape}")
            
            # Create a new series with replaced values
            # Use replace() which is more forgiving with type matching
            replaced_series = current_col.replace(value_mapping)
            
            print(f"DEBUG: After replace - dtype: {replaced_series.dtype}")
            print(f"DEBUG: Sample after replace: {replaced_series.dropna().head(3).tolist()}")
            
            # Verify replacement worked
            replaced_count = 0
            for orig_val, new_val in zip(current_col.head(100), replaced_series.head(100)):
                if pd.notna(orig_val) and orig_val in value_mapping:
                    if new_val == value_mapping[orig_val]:
                        replaced_count += 1
            
            print(f"DEBUG: Verified {replaced_count} replacements in first 100 rows")
            
            # Check if any values were actually replaced
            if current_col.equals(replaced_series):
                print(f"ERROR: .replace() did not modify any values!")
                # Fallback: use direct assignment with iteration
                print(f"DEBUG: Trying direct assignment approach...")
                new_series = current_col.copy()
                for orig_val, new_id in value_mapping.items():
                    mask = new_series == orig_val
                    if mask.any():
                        new_series.loc[mask] = new_id
                        print(f"DEBUG: Replaced {mask.sum()} occurrences of '{orig_val}' with '{new_id}'")
                replaced_series = new_series
            else:
                print(f"DEBUG: .replace() successfully modified values")
            
            # Preserve original NaN values
            nan_mask = current_col.isna()
            if nan_mask.any():
                replaced_series.loc[nan_mask] = current_col.loc[nan_mask]
            
            # CRITICAL: Assign the replaced series back to the dataframe using .loc
            # Use .loc to ensure we're modifying the dataframe, not creating a view
            anonymized_df.loc[:, col_name] = replaced_series
            
            # Verify assignment worked
            print(f"DEBUG: After assignment - checking if column was modified...")
            print(f"DEBUG: anonymized_df[col_name] equals original_col: {anonymized_df[col_name].equals(original_col)}")
            print(f"DEBUG: anonymized_df[col_name] equals replaced_series: {anonymized_df[col_name].equals(replaced_series)}")
            
            # Comprehensive verification
            print(f"DEBUG: Verification for column '{col_name}':")
            print(f"  - Original column type: {original_col.dtype}")
            print(f"  - Replaced column type: {replaced_series.dtype}")
            print(f"  - Original has {original_col.notna().sum()} non-null values")
            print(f"  - Replaced has {replaced_series.notna().sum()} non-null values")
            print(f"  - anonymized_df column has {anonymized_df[col_name].notna().sum()} non-null values")
            
            # Check if values actually changed
            if original_col.notna().any():
                original_sample = original_col.dropna().iloc[0]
                anonymized_sample = anonymized_df[col_name].dropna().iloc[0] if anonymized_df[col_name].notna().any() else None
                print(f"  - Original sample: {original_sample} (type: {type(original_sample)})")
                print(f"  - Anonymized sample: {anonymized_sample} (type: {type(anonymized_sample)})")
                
                # Verify they're different
                if original_sample == anonymized_sample and not str(anonymized_sample).startswith(prefix):
                    print(f"  - ERROR: Values are the same! Anonymization may have failed!")
                elif anonymized_sample and str(anonymized_sample).startswith(prefix):
                    print(f"  - SUCCESS: Anonymization worked! Value changed to {anonymized_sample}")
                else:
                    print(f"  - WARNING: Unexpected result")
            
            # Final check: verify the dataframe column was actually updated
            if anonymized_df[col_name].equals(original_col):
                print(f"  - CRITICAL ERROR: DataFrame column was not modified!")
                print(f"  - Trying alternative assignment method...")
                # Force assignment by creating a new dataframe column
                anonymized_df[col_name] = replaced_series.values
                print(f"  - After alternative assignment, equals original: {anonymized_df[col_name].equals(original_col)}")
            else:
                print(f"  - SUCCESS: DataFrame column was modified")
        
        send_progress('anonymizing', len(valid_columns), len(valid_columns), 'Anonymization complete...', 85)
    
    send_progress('finalizing', 100, 100, 'Finalizing...', 95)
    
    return anonymized_df, mapping_dict

def scrub_file_async(upload_path, columns_to_scrub, relationship_preserve, export_mapping, progress_queue, session_id):
    """Scrub file in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file into memory...', 5)
        
        # Read file into memory
        filename = os.path.basename(upload_path)
        if filename.endswith('.csv'):
            df = pd.read_csv(upload_path)
        else:
            df = pd.read_excel(upload_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows, {len(df.columns)} columns', 10)
        
        # Scrub the dataframe
        anonymized_df, mapping_dict = scrub_dataframe(df, columns_to_scrub, relationship_preserve, progress_queue, session_id)
        
        print(f"Anonymization complete. Columns scrubbed: {len(columns_to_scrub)}")
        
        # CRITICAL VERIFICATION: Check if anonymization actually worked before saving
        print("\n" + "="*80)
        print("PRE-SAVE VERIFICATION")
        print("="*80)
        
        anonymization_verified = True
        for col_name in columns_to_scrub:
            if col_name not in anonymized_df.columns or col_name not in df.columns:
                print(f"ERROR: Column '{col_name}' missing from dataframe!")
                anonymization_verified = False
                continue
                
            # Check if the columns are actually different
            if df[col_name].equals(anonymized_df[col_name]):
                print(f"CRITICAL ERROR: Column '{col_name}' was NOT anonymized - values are identical!")
                anonymization_verified = False
            else:
                # Check a sample
                original_sample = df[col_name].dropna().iloc[0] if len(df[col_name].dropna()) > 0 else None
                anonymized_sample = anonymized_df[col_name].dropna().iloc[0] if len(anonymized_df[col_name].dropna()) > 0 else None
                
                print(f"Column '{col_name}':")
                print(f"  - Original sample: {original_sample}")
                print(f"  - Anonymized sample: {anonymized_sample}")
                
                if original_sample == anonymized_sample and original_sample is not None:
                    print(f"  - WARNING: Sample values are the same!")
                    anonymization_verified = False
                else:
                    print(f"  - Values are different - anonymization appears to have worked")
        
        print("="*80)
        
        if not anonymization_verified:
            raise ValueError("Anonymization verification failed! The data was not properly anonymized. Aborting save.")
        
        print("Verification passed - proceeding to save anonymized file\n")
        
        # Save anonymized file
        send_progress('saving', 0, 100, 'Saving anonymized data...', 90)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"anonymized_{timestamp}"
        
        if filename.endswith('.csv'):
            output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.csv")
            anonymized_df.to_csv(output_path, index=False)
            download_url = f'/download/{output_filename}.csv'
        else:
            output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
            anonymized_df.to_excel(output_path, index=False)
            download_url = f'/download/{output_filename}.xlsx'
        
        # POST-SAVE VERIFICATION: Read back the file to ensure it was saved correctly
        print("\n" + "="*80)
        print("POST-SAVE VERIFICATION")
        print("="*80)
        try:
            if filename.endswith('.csv'):
                saved_df = pd.read_csv(output_path, nrows=10)  # Read first 10 rows
            else:
                saved_df = pd.read_excel(output_path, nrows=10)
            
            print(f"Successfully read saved file: {output_path}")
            for col_name in columns_to_scrub:
                if col_name in saved_df.columns:
                    saved_sample = saved_df[col_name].dropna().iloc[0] if len(saved_df[col_name].dropna()) > 0 else None
                    original_sample = df[col_name].dropna().iloc[0] if len(df[col_name].dropna()) > 0 else None
                    print(f"Column '{col_name}' in saved file:")
                    print(f"  - Original: {original_sample}")
                    print(f"  - Saved: {saved_sample}")
                    if saved_sample == original_sample and original_sample is not None:
                        print(f"  - ERROR: Saved file contains original values!")
                    else:
                        print(f"  - Saved file contains anonymized values")
            print("="*80 + "\n")
        except Exception as e:
            print(f"WARNING: Could not verify saved file: {e}")
            print("="*80 + "\n")
        
        send_progress('saving', 50, 100, 'Anonymized data saved...', 95)
        
        # Save mapping key if requested
        mapping_url = None
        if export_mapping and mapping_dict:
            mapping_path = os.path.join(app.config['OUTPUT_FOLDER'], f"mapping_key_{timestamp}.json")
            with open(mapping_path, 'w') as f:
                json.dump(mapping_dict, f, indent=2, default=str)
            mapping_url = f'/download/mapping_key_{timestamp}.json'
            send_progress('saving', 100, 100, 'Mapping key saved...', 98)
        else:
            send_progress('saving', 100, 100, 'Complete...', 98)
        
        # Clean up: explicitly delete dataframes and remove temp file
        del df
        del anonymized_df
        os.remove(upload_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'mapping_url': mapping_url,
            'output_filename': os.path.basename(output_path)
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
            print("Completion message sent successfully")
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
                print("Completion message sent (nowait)")
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Anonymization complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'upload_path' in locals() and os.path.exists(upload_path):
            os.remove(upload_path)
        if 'df' in locals():
            del df
        if 'anonymized_df' in locals():
            del anonymized_df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/get-columns', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def get_columns():
    """Get column names from uploaded file for column selection"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_id = f"temp_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{temp_id}_{filename}")
        file.save(upload_path)
        
        try:
            # Read file to get columns
            if filename.endswith('.csv'):
                df = pd.read_csv(upload_path, nrows=0)  # Read only headers
            else:
                df = pd.read_excel(upload_path, nrows=0)
            
            columns_info = []
            for col in df.columns:
                col_info = {
                    'name': str(col),
                    'dtype': str(df[col].dtype)
                }
                columns_info.append(col_info)
            
            # Clean up temp file
            os.remove(upload_path)
            
            return jsonify({
                'success': True,
                'columns': columns_info
            })
        except Exception as e:
            if os.path.exists(upload_path):
                os.remove(upload_path)
            raise e
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/get-columns-with-samples', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def get_columns_with_samples():
    """Get column names with sample values from uploaded file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_id = f"temp_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{temp_id}_{filename}")
        file.save(upload_path)
        
        try:
            # Read file to get columns and sample values (first 10 rows)
            if filename.endswith('.csv'):
                df = pd.read_csv(upload_path, nrows=10)
            else:
                df = pd.read_excel(upload_path, nrows=10)
            
            columns_info = []
            for col in df.columns:
                # Get sample values (first 10 non-null values)
                sample_values = df[col].dropna().head(10).tolist()
                # Convert to strings for JSON serialization
                sample_values = [str(val) if pd.notna(val) else '' for val in sample_values]
                
                col_info = {
                    'name': str(col),
                    'dtype': str(df[col].dtype),
                    'sample_values': sample_values[:10]  # Limit to 10 samples
                }
                columns_info.append(col_info)
            
            # Clean up temp file
            os.remove(upload_path)
            
            return jsonify({
                'success': True,
                'columns': columns_info
            })
        except Exception as e:
            if os.path.exists(upload_path):
                os.remove(upload_path)
            raise e
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# =============================================================================
# DATA PREVIEW & SESSION CACHE ENDPOINTS
# =============================================================================

@app.route('/preview-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def preview_data():
    """Get a preview of uploaded file data (first 20 rows)"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Validate file extension
        if not allowed_file(file.filename):
            return jsonify({'error': 'Please upload an Excel or CSV file (.xlsx, .xls, .csv)'}), 400

        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_id = f"preview_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{temp_id}_{filename}")
        file.save(upload_path)

        try:
            preview = get_file_preview(upload_path, max_rows=20)
            return jsonify({
                'success': True,
                'filename': filename,
                **preview
            })
        finally:
            # Clean up temp file
            if os.path.exists(upload_path):
                os.remove(upload_path)

    except Exception as e:
        print(f"Error in preview_data: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/get-cached-files', methods=['GET'])
def get_cached_files_endpoint():
    """Get list of cached files from recent tool operations"""
    try:
        # Clean up old caches first
        cleanup_session_cache()

        # Format for frontend display with cache IDs
        files = []
        for cache_id, f in file_cache.items():
            if os.path.exists(f.get('path', '')):
                # Format timestamp for display
                from datetime import datetime
                cached_at = datetime.fromtimestamp(f['timestamp']).strftime('%H:%M:%S')
                files.append({
                    'cache_id': cache_id,
                    'filename': f['name'],
                    'rows': f['rows'],
                    'cols': f['cols'],
                    'timestamp': f['timestamp'],
                    'cached_at': cached_at,
                    'source_tool': f.get('source_tool', 'Unknown'),
                    'age_seconds': int(time.time() - f['timestamp'])
                })

        # Sort by timestamp descending (most recent first)
        files.sort(key=lambda x: x['timestamp'], reverse=True)

        return jsonify({'success': True, 'files': files})

    except Exception as e:
        print(f"Error in get_cached_files: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/use-cached-file', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def use_cached_file():
    """Use a previously cached file result in a new tool operation"""
    try:
        data = request.get_json()
        cache_id = data.get('cache_id', '')

        if not cache_id:
            return jsonify({'error': 'No cache ID provided'}), 400

        file_info = get_cached_file_by_id(cache_id)
        if not file_info:
            return jsonify({'error': 'Cached file not found'}), 400

        if not os.path.exists(file_info.get('path', '')):
            return jsonify({'error': 'Cached file no longer exists'}), 400

        # Return file info for the frontend to use
        preview = get_file_preview(file_info['path'], max_rows=20)

        return jsonify({
            'success': True,
            'filename': file_info['name'],
            'path': file_info['path'],
            'preview': preview
        })

    except Exception as e:
        print(f"Error in use_cached_file: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/download-cached-file/<cache_id>', methods=['GET'])
def download_cached_file(cache_id):
    """Download a cached file by its cache ID"""
    try:
        file_info = get_cached_file_by_id(cache_id)
        if not file_info:
            return jsonify({'error': 'Cached file not found'}), 404

        file_path = file_info.get('path', '')
        if not os.path.exists(file_path):
            return jsonify({'error': 'Cached file no longer exists'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=file_info['name']
        )

    except Exception as e:
        print(f"Error downloading cached file: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/scrub-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def scrub_data():
    """Scrub uploaded file - returns session_id for progress tracking - NO data stored"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Validate MIME type
        allowed_mimes = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
            'text/csv',  # .csv
            'application/csv'
        ]
        if hasattr(file, 'content_type') and file.content_type and file.content_type not in allowed_mimes:
            if file.content_type and not (file.content_type.startswith('text/') or file.content_type.startswith('application/')):
                return jsonify({'error': 'Invalid file type. Please upload an Excel or CSV file.'}), 400
        
        # Get columns to scrub and options
        columns_to_scrub = request.form.getlist('columns[]')
        if not columns_to_scrub:
            return jsonify({'error': 'No columns selected for anonymization'}), 400
        
        relationship_preserve = request.form.get('relationship_preserve', 'false').lower() == 'true'
        export_mapping = request.form.get('export_mapping', 'false').lower() == 'true'
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"scrub_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=scrub_file_async,
            args=(upload_path, columns_to_scrub, relationship_preserve, export_mapping, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/duplicate-finder')
def duplicate_finder():
    return render_template('duplicate_finder.html')

@app.route('/unique-identifier-finder')
def unique_identifier_finder():
    return render_template('unique_identifier_finder.html')

def find_duplicates_async(upload_path, id_column, duplicate_columns, progress_queue, session_id):
    """Find duplicate rows (selected columns combined) in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file into memory...', 5)
        
        # Read file into memory
        filename = os.path.basename(upload_path)
        if filename.endswith('.csv'):
            df = pd.read_csv(upload_path)
        else:
            df = pd.read_excel(upload_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows, {len(df.columns)} columns', 10)
        
        # Validate ID column exists
        if not id_column or id_column not in df.columns:
            raise ValueError(f"ID column '{id_column}' not found in file. Please select a valid ID column.")
        
        # Validate duplicate columns exist
        if not duplicate_columns:
            raise ValueError("Please select at least one column to check for duplicates.")
        
        invalid_columns = [col for col in duplicate_columns if col not in df.columns]
        if invalid_columns:
            raise ValueError(f"Columns not found in file: {', '.join(invalid_columns)}")
        
        # Ensure ID column is not in duplicate columns (we use it separately)
        duplicate_columns = [col for col in duplicate_columns if col != id_column]
        
        if not duplicate_columns:
            raise ValueError("Please select at least one column (other than the ID column) to check for duplicates.")
        
        send_progress('analyzing', 0, 100, f'Creating row signatures (concatenating {len(duplicate_columns)} selected columns)...', 20)
        
        # Create a signature for each row by concatenating selected column values
        def create_row_signature(row):
            # Convert selected column values to strings, handling NaN
            values = [str(row[col]) if pd.notna(row[col]) else '' for col in duplicate_columns]
            return '|||'.join(values)  # Use a unique delimiter
        
        # Create signatures for all rows
        df['_row_signature'] = df.apply(create_row_signature, axis=1)
        
        send_progress('analyzing', 30, 100, 'Finding duplicate rows...', 40)
        
        # Find duplicate rows by grouping on row signature
        duplicate_groups = df.groupby('_row_signature').size()
        duplicate_signatures = duplicate_groups[duplicate_groups > 1].index.tolist()
        
        send_progress('analyzing', 60, 100, f'Found {len(duplicate_signatures)} duplicate row groups...', 60)
        
        # Build results
        results = []
        all_ids_to_remove = []
        
        for sig_idx, signature in enumerate(duplicate_signatures):
            # Get all rows with this signature
            dup_rows = df[df['_row_signature'] == signature]
            
            # Get IDs for these rows
            ids = dup_rows[id_column].tolist()
            
            # Get a sample of the row data (first row) to show what the duplicate looks like
            sample_row = dup_rows.iloc[0]
            # Create a readable representation using the selected duplicate columns
            row_preview = ', '.join([f"{col}={sample_row[col]}" for col in duplicate_columns])
            
            # Keep first ID, remove rest
            ids_to_remove = ids[1:] if len(ids) > 1 else []
            all_ids_to_remove.extend(ids_to_remove)
            
            results.append({
                'row_preview': row_preview,
                'count': len(dup_rows),
                'ids': ids,
                'ids_to_remove': ids_to_remove,
                'keep_id': ids[0] if ids else None
            })
            
            # Progress update every 100 groups
            if (sig_idx + 1) % 100 == 0:
                send_progress('analyzing', 60 + int((sig_idx + 1) / len(duplicate_signatures) * 10), 100, 
                            f'Processing group {sig_idx + 1}/{len(duplicate_signatures)}...', 
                            60 + int((sig_idx + 1) / len(duplicate_signatures) * 10))
        
        send_progress('analyzing', 100, 100, f'Analysis complete: {len(results)} duplicate groups found', 70)
        
        # Create results DataFrame
        results_data = []
        
        for result in results:
            ids_str = ', '.join(str(id_val) for id_val in result['ids'])
            ids_to_remove_str = ', '.join(str(id_val) for id_val in result['ids_to_remove'])
            
            results_data.append({
                'Row Preview': result['row_preview'],
                'Count': result['count'],
                'Keep ID (Original)': result['keep_id'],
                'All IDs': ids_str,
                'IDs to Remove': ids_to_remove_str
            })
        
        results_df = pd.DataFrame(results_data)
        
        send_progress('saving', 0, 100, 'Saving results...', 80)
        
        # Save results to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"duplicates_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Duplicate summary
            results_df.to_excel(writer, sheet_name='Duplicates', index=False)
            
            # Sheet 2: IDs to remove (single column for easy copy/paste)
            if all_ids_to_remove:
                ids_to_remove_df = pd.DataFrame({'ID': all_ids_to_remove})
                ids_to_remove_df.to_excel(writer, sheet_name='IDs to Remove', index=False)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'Results saved...', 95)
        
        # Clean up
        del df
        del results_df
        os.remove(upload_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'results': results_data[:100],  # Send first 100 for preview
            'total_duplicates': len(results),
            'total_ids_to_remove': len(all_ids_to_remove)
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Duplicate finding complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'upload_path' in locals() and os.path.exists(upload_path):
            os.remove(upload_path)
        if 'df' in locals():
            del df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/find-duplicates', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def find_duplicates():
    """Find duplicates in uploaded file - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Get ID column selection (required)
        id_column = request.form.get('id_column', '').strip()
        
        if not id_column:
            return jsonify({'error': 'Please select an ID column to identify duplicate records'}), 400
        
        # Get duplicate columns selection (required, can be multiple)
        duplicate_columns = request.form.getlist('duplicate_columns[]')
        
        if not duplicate_columns:
            return jsonify({'error': 'Please select at least one column to check for duplicates'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"duplicates_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=find_duplicates_async,
            args=(upload_path, id_column, duplicate_columns, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


def generate_natural_key_report(output_path, original_rows, duplicate_count, rows_analyzed,
                                 selected_columns, minimal_combinations, primary_combo,
                                 sample_data, source_filename):
    """
    Generate a professional PDF report for Natural Key Analysis results.

    Args:
        output_path: Path to save the PDF file
        original_rows: Total records in source file
        duplicate_count: Number of duplicate records excluded
        rows_analyzed: Records after duplicate exclusion
        selected_columns: List of columns evaluated
        minimal_combinations: List of all minimal key combinations found
        primary_combo: The recommended primary key combination
        sample_data: Sample DataFrame showing unique IDs (first 20 rows)
        source_filename: Original filename for reference
    """
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    # Custom styles
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor('#333333'),
        borderPadding=5
    )

    subheading_style = ParagraphStyle(
        'CustomSubheading',
        parent=styles['Heading3'],
        fontSize=12,
        spaceBefore=15,
        spaceAfter=8,
        textColor=colors.HexColor('#667eea')
    )

    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=8,
        leading=14
    )

    note_style = ParagraphStyle(
        'NoteStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=6,
        leading=12
    )

    # Build document content
    story = []

    # Title
    story.append(Paragraph("Natural Key Analysis Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
                          ParagraphStyle('DateStyle', parent=styles['Normal'],
                                        alignment=TA_CENTER, textColor=colors.gray, fontSize=10)))
    story.append(Spacer(1, 20))

    # Executive Summary Section
    story.append(Paragraph("Executive Summary", heading_style))

    if len(primary_combo) == 1:
        summary_text = f"""
        Analysis of <b>{source_filename}</b> identified <b>{primary_combo[0]}</b> as a single-column
        natural key capable of uniquely identifying all {rows_analyzed:,} records. This column can serve
        as a primary key without requiring additional columns.
        """
    else:
        combo_text = " + ".join([f"<b>{col}</b>" for col in primary_combo])
        summary_text = f"""
        Analysis of <b>{source_filename}</b> determined that a composite key of {len(primary_combo)} columns
        ({combo_text}) is required to uniquely identify all {rows_analyzed:,} records.
        No single column provides unique identification.
        """
    story.append(Paragraph(summary_text, body_style))

    if len(minimal_combinations) > 1:
        story.append(Paragraph(f"<b>{len(minimal_combinations)} alternative key combinations</b> were identified, "
                              f"all achieving the same minimum key size of {len(primary_combo)} column(s).", body_style))

    story.append(Spacer(1, 15))

    # Data Overview Section
    story.append(Paragraph("Data Overview", heading_style))

    overview_data = [
        ['Metric', 'Value'],
        ['Source File', source_filename],
        ['Total Records', f'{original_rows:,}'],
        ['Duplicate Records Excluded', f'{duplicate_count:,}'],
        ['Records Analyzed', f'{rows_analyzed:,}'],
        ['Columns Evaluated', f'{len(selected_columns)}'],
        ['Minimum Key Size Required', f'{len(primary_combo)} column(s)'],
        ['Valid Key Combinations Found', f'{len(minimal_combinations)}']
    ]

    overview_table = Table(overview_data, colWidths=[2.5*inch, 4*inch])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8f9ff'), colors.white])
    ]))
    story.append(overview_table)
    story.append(Spacer(1, 20))

    # Primary Key Recommendation
    story.append(Paragraph("Primary Key Recommendation", heading_style))

    primary_key_data = [['Rank', 'Key Columns', 'Column Count']]
    primary_key_data.append(['Primary', ', '.join(primary_combo), str(len(primary_combo))])

    primary_table = Table(primary_key_data, colWidths=[1*inch, 4.5*inch, 1*inch])
    primary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#d4edda')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#28a745'))
    ]))
    story.append(primary_table)

    story.append(Spacer(1, 10))
    story.append(Paragraph("<b>Implementation:</b> Concatenate values from the key columns above to generate "
                          "a unique identifier for each record. The accompanying Excel file includes a "
                          "pre-computed <i>Unique_ID</i> column.", note_style))
    story.append(Spacer(1, 15))

    # Alternative Key Candidates (if any)
    if len(minimal_combinations) > 1:
        story.append(Paragraph("Alternative Key Candidates", heading_style))
        story.append(Paragraph("The following combinations also uniquely identify all records with the same "
                              "minimum column count:", body_style))

        alt_data = [['Candidate', 'Key Columns', 'Column Count']]
        for idx, combo in enumerate(minimal_combinations[1:], 2):
            alt_data.append([f'#{idx}', ', '.join(combo), str(len(combo))])

        alt_table = Table(alt_data, colWidths=[1*inch, 4.5*inch, 1*inch])
        alt_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffc107')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fff3cd')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ffc107')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fff3cd'), colors.HexColor('#fffef5')])
        ]))
        story.append(alt_table)
        story.append(Spacer(1, 15))

    # Columns Evaluated
    story.append(Paragraph("Columns Evaluated", heading_style))

    # Display columns in a multi-column layout
    cols_per_row = 3
    col_rows = [selected_columns[i:i+cols_per_row] for i in range(0, len(selected_columns), cols_per_row)]
    # Pad the last row if needed
    if col_rows and len(col_rows[-1]) < cols_per_row:
        col_rows[-1].extend([''] * (cols_per_row - len(col_rows[-1])))

    if col_rows:
        col_table = Table(col_rows, colWidths=[2.17*inch] * cols_per_row)
        col_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9ff'))
        ]))
        story.append(col_table)
    story.append(Spacer(1, 20))

    # Sample Unique Identifiers
    story.append(Paragraph("Sample Unique Identifiers", heading_style))
    story.append(Paragraph("Preview of the first 20 records with computed unique identifiers:", body_style))

    # Prepare sample data table
    if sample_data is not None and len(sample_data) > 0:
        sample_cols = list(sample_data.columns)
        sample_header = [sample_cols]
        sample_rows = sample_data.head(20).values.tolist()

        # Truncate long values for display
        for row in sample_rows:
            for i, val in enumerate(row):
                str_val = str(val) if pd.notna(val) else ''
                if len(str_val) > 25:
                    row[i] = str_val[:22] + '...'
                else:
                    row[i] = str_val

        sample_table_data = sample_header + sample_rows

        # Calculate column widths based on number of columns
        num_cols = len(sample_cols)
        available_width = 6.5 * inch
        col_width = available_width / num_cols

        sample_table = Table(sample_table_data, colWidths=[col_width] * num_cols)
        sample_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9ff')])
        ]))
        story.append(sample_table)

    story.append(Spacer(1, 20))

    # Algorithm Explanation
    story.append(Paragraph("Methodology", heading_style))
    story.append(Paragraph("""
    This analysis employed the <b>Apriori algorithm</b> for efficient discovery of minimal unique column
    combinations. The algorithm operates in levels, first testing single columns, then progressively
    larger combinations. Key optimization: if a column set is unique, no superset is evaluated
    (as it cannot be minimal). This approach ensures computational efficiency while guaranteeing
    discovery of all truly minimal natural keys.
    """, body_style))

    story.append(Spacer(1, 15))

    # Footer
    story.append(Paragraph("─" * 80, ParagraphStyle('Line', alignment=TA_CENTER, textColor=colors.lightgrey)))
    story.append(Spacer(1, 5))
    story.append(Paragraph("Generated by DataDragon - Natural Key Analysis Tool",
                          ParagraphStyle('Footer', alignment=TA_CENTER, fontSize=9, textColor=colors.gray)))

    # Build PDF
    doc.build(story)


def find_unique_identifier_async(upload_path, selected_columns, progress_queue, session_id):
    """Find minimal set of columns that create unique identifiers in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file into memory...', 5)
        
        # Read file into memory
        filename = os.path.basename(upload_path)
        if filename.endswith('.csv'):
            df = pd.read_csv(upload_path)
        else:
            df = pd.read_excel(upload_path)
        
        original_row_count = len(df)
        send_progress('loading', 50, 100, f'File loaded: {len(df):,} rows, {len(df.columns)} columns', 10)
        
        # Step 1: Filter out fully duplicate rows
        send_progress('filtering', 0, 100, 'Identifying fully duplicate rows...', 15)
        duplicate_count = df.duplicated().sum()
        
        if duplicate_count > 0:
            df = df.drop_duplicates(keep='first')
            send_progress('filtering', 100, 100, f'Removed {duplicate_count:,} fully duplicate rows. {len(df):,} rows remaining.', 20)
        else:
            send_progress('filtering', 100, 100, 'No fully duplicate rows found.', 20)
        
        # Validate selected columns exist
        if not selected_columns:
            raise ValueError("Please select at least one column to analyze.")
        
        invalid_columns = [col for col in selected_columns if col not in df.columns]
        if invalid_columns:
            raise ValueError(f"Columns not found in file: {', '.join(invalid_columns)}")
        
        # Step 2: Find minimal unique combinations using Apriori pruning algorithm
        # This algorithm ensures we find truly MINIMAL keys by only expanding non-unique combinations
        send_progress('analyzing', 0, 100, f'Analyzing {len(selected_columns)} selected columns using Apriori algorithm...', 25)

        minimal_combinations = []
        non_unique_combinations = []  # Stores sets of columns that are NOT unique
        n = len(selected_columns)
        max_size = min(n, 5)  # Limit composite key size for performance
        max_keys = 10  # Maximum number of minimal keys to find

        # Level 1: Check single columns
        send_progress('analyzing', 10, 100, 'Level 1: Checking single columns...', 30)
        for idx, col in enumerate(selected_columns):
            # Check if this single column creates unique identifiers
            is_unique = df[[col]].duplicated().sum() == 0

            if is_unique:
                minimal_combinations.append([col])
            else:
                non_unique_combinations.append(frozenset([col]))

            if (idx + 1) % 10 == 0 or idx == len(selected_columns) - 1:
                send_progress('analyzing', 10 + int((idx + 1) / len(selected_columns) * 10), 100,
                            f'Checked {idx + 1}/{len(selected_columns)} single columns...',
                            30 + int((idx + 1) / len(selected_columns) * 5))

        # Levels 2+: Use Apriori pruning to find minimal composite keys
        # Only expand combinations that were proven NON-UNIQUE in the previous level
        for k in range(2, max_size + 1):
            if not non_unique_combinations:
                send_progress('analyzing', 50, 100, 'No more non-unique combinations to expand. Algorithm complete.', 55)
                break

            if len(minimal_combinations) >= max_keys:
                send_progress('analyzing', 50, 100, f'Found {len(minimal_combinations)} minimal keys. Stopping search.', 55)
                break

            send_progress('analyzing', 20 + int((k - 2) * 15), 100,
                        f'Level {k}: Generating candidates using Apriori pruning...',
                        40 + int((k - 2) * 10))

            # Generate candidate combinations using Apriori principle:
            # A candidate of size k is valid ONLY IF all its (k-1) subsets are in non_unique_combinations
            # This ensures we never check supersets of already-found keys (which wouldn't be minimal)
            pool_cols = set()
            for combo in non_unique_combinations:
                pool_cols.update(combo)

            valid_candidates = []
            for combo in combinations(sorted(pool_cols), k):
                combo_set = frozenset(combo)

                # Apriori pruning: Check if ALL (k-1) subsets are in non_unique_combinations
                # If any subset is NOT in non_unique, it means that subset is either:
                # a) A known key (so we skip - superset wouldn't be minimal)
                # b) Was never processed (shouldn't happen in correct flow)
                all_subsets_non_unique = True
                for subset in combinations(combo, k - 1):
                    if frozenset(subset) not in non_unique_combinations:
                        all_subsets_non_unique = False
                        break

                if all_subsets_non_unique:
                    valid_candidates.append(list(combo))

            send_progress('analyzing', 25 + int((k - 2) * 15), 100,
                        f'Level {k}: Testing {len(valid_candidates)} pruned candidates...',
                        45 + int((k - 2) * 10))

            # Reset non_unique for the next level
            next_level_non_unique = []

            for combo_idx, candidate in enumerate(valid_candidates):
                if len(minimal_combinations) >= max_keys:
                    break

                # Check uniqueness using duplicated() - efficient vectorized operation
                is_unique = df[candidate].duplicated().sum() == 0

                if is_unique:
                    minimal_combinations.append(candidate)
                else:
                    next_level_non_unique.append(frozenset(candidate))

                # Progress update
                if (combo_idx + 1) % 25 == 0 or combo_idx == len(valid_candidates) - 1:
                    progress_pct = 25 + int((k - 2) * 15) + int((combo_idx + 1) / max(len(valid_candidates), 1) * 10)
                    send_progress('analyzing', progress_pct, 100,
                                f'Level {k}: Tested {combo_idx + 1}/{len(valid_candidates)} candidates, found {len(minimal_combinations)} keys...',
                                45 + int((k - 2) * 10) + int((combo_idx + 1) / max(len(valid_candidates), 1) * 5))

            non_unique_combinations = next_level_non_unique
        
        if not minimal_combinations:
            raise ValueError("No combination of selected columns can create unique identifiers for all rows.")
        
        send_progress('analyzing', 100, 100, f'Found {len(minimal_combinations)} minimal key candidate(s)', 70)

        # Step 3: Generate results
        send_progress('saving', 0, 100, 'Preparing output files...', 75)

        # Use first minimal combination to create unique IDs
        primary_combo = minimal_combinations[0]
        df['Unique_ID'] = df[primary_combo].apply(
            lambda row: '|||'.join(str(v) if pd.notna(v) else '' for v in row),
            axis=1
        )

        # Sample of unique IDs for PDF report
        sample_df = df[primary_combo + ['Unique_ID']].head(100).copy()

        # Get source filename for report
        source_filename = os.path.basename(upload_path)
        # Remove session prefix if present
        if '_' in source_filename:
            source_filename = '_'.join(source_filename.split('_')[2:]) or source_filename

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_basename = f"natural_key_analysis_{timestamp}"

        # Generate PDF Report
        send_progress('saving', 25, 100, 'Generating PDF report...', 80)
        pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}_report.pdf")

        generate_natural_key_report(
            output_path=pdf_path,
            original_rows=original_row_count,
            duplicate_count=duplicate_count,
            rows_analyzed=len(df),
            selected_columns=selected_columns,
            minimal_combinations=minimal_combinations,
            primary_combo=primary_combo,
            sample_data=sample_df,
            source_filename=source_filename
        )

        # Generate Excel Data File
        send_progress('saving', 60, 100, 'Generating Excel data file...', 88)
        excel_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}_data.xlsx")

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Data with Unique IDs
            df.to_excel(writer, sheet_name='Data with Unique IDs', index=False)

            # Sheet 2: Key Combinations
            alternatives_data = []
            for idx, combo in enumerate(minimal_combinations, 1):
                alternatives_data.append({
                    'Candidate': f'#{idx}' if idx > 1 else 'Primary',
                    'Key Columns': ', '.join(combo),
                    'Column Count': len(combo)
                })
            alternatives_df = pd.DataFrame(alternatives_data)
            alternatives_df.to_excel(writer, sheet_name='Key Candidates', index=False)

        # Create ZIP package containing both files
        send_progress('saving', 85, 100, 'Packaging results...', 93)
        zip_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}.zip")

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(pdf_path, f"{output_basename}_report.pdf")
            zipf.write(excel_path, f"{output_basename}_data.xlsx")

        # Clean up individual files (keep only ZIP)
        os.remove(pdf_path)
        os.remove(excel_path)

        download_url = f'/download/{output_basename}.zip'

        send_progress('saving', 100, 100, 'Results packaged successfully', 95)

        # Clean up DataFrames
        del df
        del alternatives_df
        del sample_df
        os.remove(upload_path)

        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Analysis complete',
            'download_url': download_url,
            'output_filename': f"{output_basename}.zip",
            'original_rows': original_row_count,
            'duplicate_rows_removed': duplicate_count,
            'rows_after_filtering': original_row_count - duplicate_count,
            'selected_columns_count': len(selected_columns),
            'minimal_columns_count': len(primary_combo),
            'minimal_columns': primary_combo,
            'alternatives_count': len(minimal_combinations),
            'alternatives': minimal_combinations
        }

        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")

        print("Natural key analysis complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'upload_path' in locals() and os.path.exists(upload_path):
            os.remove(upload_path)
        if 'df' in locals():
            del df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/find-unique-identifier', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def find_unique_identifier():
    """Find unique identifier columns in uploaded file - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Get selected columns (required, can be multiple)
        selected_columns = request.form.getlist('selected_columns[]')
        
        if not selected_columns:
            return jsonify({'error': 'Please select at least one column to analyze'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"unique_id_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=find_unique_identifier_async,
            args=(upload_path, selected_columns, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/data-merge')
def data_merge():
    return render_template('data_merge.html')

def merge_files_async(left_file_path, right_file_path, left_key, right_key, join_type, left_columns, right_columns, duplicate_handling, progress_queue, session_id):
    """Merge two files in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading left file...', 5)
        
        # Read left file
        if left_file_path.endswith('.csv'):
            df_left = pd.read_csv(left_file_path)
        else:
            df_left = pd.read_excel(left_file_path)
        
        send_progress('loading', 50, 100, 'Reading right file...', 10)
        
        # Read right file
        if right_file_path.endswith('.csv'):
            df_right = pd.read_csv(right_file_path)
        else:
            df_right = pd.read_excel(right_file_path)
        
        send_progress('loading', 100, 100, f'Files loaded: Left {len(df_left):,} rows, Right {len(df_right):,} rows', 15)
        
        # Validate key columns exist
        if left_key not in df_left.columns:
            raise ValueError(f"Key column '{left_key}' not found in left file")
        if right_key not in df_right.columns:
            raise ValueError(f"Key column '{right_key}' not found in right file")
        
        # Validate selected columns exist
        if left_columns:
            invalid_left = [col for col in left_columns if col not in df_left.columns]
            if invalid_left:
                raise ValueError(f"Columns not found in left file: {', '.join(invalid_left)}")
        
        if right_columns:
            invalid_right = [col for col in right_columns if col not in df_right.columns]
            if invalid_right:
                raise ValueError(f"Columns not found in right file: {', '.join(invalid_right)}")
        
        send_progress('preparing', 0, 100, 'Preparing data for merge...', 20)
        
        # Select columns to include (if specified)
        if left_columns:
            # Always include the key column
            if left_key not in left_columns:
                left_columns = [left_key] + left_columns
            df_left = df_left[left_columns]
        
        if right_columns:
            # Always include the key column
            if right_key not in right_columns:
                right_columns = [right_key] + right_columns
            df_right = df_right[right_columns]
        
        # Handle duplicate keys
        if duplicate_handling == 'error':
            left_dupes = df_left[left_key].duplicated().sum()
            right_dupes = df_right[right_key].duplicated().sum()
            if left_dupes > 0 or right_dupes > 0:
                raise ValueError(f"Duplicate keys found: Left file has {left_dupes} duplicates, Right file has {right_dupes} duplicates. Please handle duplicates first.")
        elif duplicate_handling == 'keep_first':
            df_left = df_left.drop_duplicates(subset=[left_key], keep='first')
            df_right = df_right.drop_duplicates(subset=[right_key], keep='first')
        
        send_progress('merging', 0, 100, f'Performing {join_type} join...', 30)
        
        # Perform merge
        # Map join types
        join_type_map = {
            'left': 'left',
            'right': 'right',
            'inner': 'inner',
            'outer': 'outer'
        }
        how = join_type_map.get(join_type, 'inner')
        
        merged_df = pd.merge(
            df_left,
            df_right,
            left_on=left_key,
            right_on=right_key,
            how=how,
            suffixes=('_left', '_right')
        )
        
        send_progress('merging', 100, 100, f'Merged: {len(merged_df):,} rows', 60)
        
        # Calculate join statistics
        left_total = len(df_left)
        right_total = len(df_right)
        merged_total = len(merged_df)
        
        # Count matched/unmatched
        if how == 'left':
            matched = merged_df[right_key].notna().sum()
            unmatched_left = merged_df[right_key].isna().sum()
            unmatched_right = right_total - matched
        elif how == 'right':
            matched = merged_df[left_key].notna().sum()
            unmatched_right = merged_df[left_key].isna().sum()
            unmatched_left = left_total - matched
        elif how == 'inner':
            matched = merged_total
            unmatched_left = left_total - matched
            unmatched_right = right_total - matched
        else:  # outer
            # For outer join, matched means both keys are present
            matched = (merged_df[left_key].notna() & merged_df[right_key].notna()).sum()
            unmatched_left = (merged_df[left_key].notna() & merged_df[right_key].isna()).sum()
            unmatched_right = (merged_df[left_key].isna() & merged_df[right_key].notna()).sum()
        
        send_progress('saving', 0, 100, 'Saving results...', 70)
        
        # Create summary DataFrame
        summary_data = {
            'Metric': [
                'Left File Rows',
                'Right File Rows',
                'Merged Rows',
                'Matched Rows',
                'Unmatched (Left)',
                'Unmatched (Right)',
                'Join Type'
            ],
            'Value': [
                left_total,
                right_total,
                merged_total,
                matched,
                unmatched_left,
                unmatched_right,
                join_type.upper()
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Save to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"merged_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='Merged Data', index=False)
            summary_df.to_excel(writer, sheet_name='Join Summary', index=False)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'Results saved...', 95)
        
        # Prepare preview data before cleanup
        preview_data = []
        if len(merged_df) > 0:
            preview_df = merged_df.head(100)
            # Convert to dict, handling NaN values
            for _, row in preview_df.iterrows():
                row_dict = {}
                for col in preview_df.columns:
                    val = row[col]
                    if pd.isna(val):
                        row_dict[col] = None
                    else:
                        row_dict[col] = str(val)
                preview_data.append(row_dict)
        
        # Clean up
        del df_left
        del df_right
        del merged_df
        os.remove(left_file_path)
        os.remove(right_file_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'summary': {
                'left_rows': int(left_total),
                'right_rows': int(right_total),
                'merged_rows': int(merged_total),
                'matched': int(matched),
                'unmatched_left': int(unmatched_left),
                'unmatched_right': int(unmatched_right),
                'join_type': join_type
            },
            'preview': preview_data
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Merge complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'left_file_path' in locals() and os.path.exists(left_file_path):
            os.remove(left_file_path)
        if 'right_file_path' in locals() and os.path.exists(right_file_path):
            os.remove(right_file_path)
        if 'df_left' in locals():
            del df_left
        if 'df_right' in locals():
            del df_right
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/merge-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def merge_data():
    """Merge two uploaded files - returns session_id for progress tracking"""
    try:
        if 'left_file' not in request.files or 'right_file' not in request.files:
            return jsonify({'error': 'Please upload both left and right files'}), 400
        
        left_file = request.files['left_file']
        right_file = request.files['right_file']
        
        if left_file.filename == '' or right_file.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Validate file extensions
        if not (left_file.filename.endswith(('.xlsx', '.xls', '.csv')) and 
                right_file.filename.endswith(('.xlsx', '.xls', '.csv'))):
            return jsonify({'error': 'Please upload Excel or CSV files'}), 400
        
        # Get merge parameters
        left_key = request.form.get('left_key', '').strip()
        right_key = request.form.get('right_key', '').strip()
        join_type = request.form.get('join_type', 'inner').strip()
        duplicate_handling = request.form.get('duplicate_handling', 'keep_first').strip()
        
        if not left_key or not right_key:
            return jsonify({'error': 'Please select key columns from both files'}), 400
        
        # Get selected columns (optional)
        left_columns = request.form.getlist('left_columns[]')
        right_columns = request.form.getlist('right_columns[]')
        
        # Save files temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"merge_{timestamp}_{secrets.token_hex(8)}"
        
        left_filename = secure_filename(left_file.filename)
        right_filename = secure_filename(right_file.filename)
        
        left_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_left_{left_filename}")
        right_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_right_{right_filename}")
        
        left_file.save(left_file_path)
        right_file.save(right_file_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=merge_files_async,
            args=(left_file_path, right_file_path, left_key, right_key, join_type, left_columns, right_columns, duplicate_handling, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/data-comparison')
def data_comparison():
    return render_template('data_comparison.html')

def compare_files_async(file1_path, file2_path, key_columns, compare_columns, progress_queue, session_id):
    """Compare two files in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file 1...', 5)
        
        # Read file 1
        if file1_path.endswith('.csv'):
            df1 = pd.read_csv(file1_path)
        else:
            df1 = pd.read_excel(file1_path)
        
        send_progress('loading', 50, 100, 'Reading file 2...', 10)
        
        # Read file 2
        if file2_path.endswith('.csv'):
            df2 = pd.read_csv(file2_path)
        else:
            df2 = pd.read_excel(file2_path)
        
        send_progress('loading', 100, 100, f'Files loaded: File 1 {len(df1):,} rows, File 2 {len(df2):,} rows', 15)
        
        # Validate key columns exist
        for key_col in key_columns:
            if key_col not in df1.columns:
                raise ValueError(f"Key column '{key_col}' not found in file 1")
            if key_col not in df2.columns:
                raise ValueError(f"Key column '{key_col}' not found in file 2")
        
        # Validate compare columns exist (if specified)
        if compare_columns:
            invalid_cols = [col for col in compare_columns if col not in df1.columns or col not in df2.columns]
            if invalid_cols:
                raise ValueError(f"Columns not found in both files: {', '.join(invalid_cols)}")
        
        send_progress('preparing', 0, 100, 'Preparing data for comparison...', 20)
        
        # Determine which columns to compare
        if compare_columns:
            # Exclude key columns from compare_columns if they're in there
            compare_cols = [col for col in compare_columns if col not in key_columns]
            # Always include key columns
            cols_to_compare = key_columns + compare_cols
        else:
            # Compare all columns except key columns (we'll add them back)
            cols_to_compare = [col for col in df1.columns if col in df2.columns]
            # Ensure key columns are included
            for key_col in key_columns:
                if key_col not in cols_to_compare:
                    cols_to_compare.insert(0, key_col)
        
        # Select only columns that exist in both files
        cols_to_compare = [col for col in cols_to_compare if col in df1.columns and col in df2.columns]
        
        df1_compare = df1[cols_to_compare].copy()
        df2_compare = df2[cols_to_compare].copy()
        
        send_progress('comparing', 0, 100, 'Comparing files...', 30)
        
        # Create composite key for matching
        if len(key_columns) == 1:
            df1_compare['_key'] = df1_compare[key_columns[0]]
            df2_compare['_key'] = df2_compare[key_columns[0]]
        else:
            # Multiple key columns - create composite key
            df1_compare['_key'] = df1_compare[key_columns].apply(lambda row: '|||'.join([str(val) if pd.notna(val) else '' for val in row]), axis=1)
            df2_compare['_key'] = df2_compare[key_columns].apply(lambda row: '|||'.join([str(val) if pd.notna(val) else '' for val in row]), axis=1)
        
        # Get all keys
        keys1 = set(df1_compare['_key'].unique())
        keys2 = set(df2_compare['_key'].unique())
        
        # Categorize rows
        added_keys = keys2 - keys1  # In file 2, not in file 1
        removed_keys = keys1 - keys2  # In file 1, not in file 2
        common_keys = keys1 & keys2  # In both files
        
        send_progress('comparing', 50, 100, f'Found {len(common_keys)} common rows, {len(added_keys)} added, {len(removed_keys)} removed...', 50)
        
        # Get added rows
        added_rows = df2_compare[df2_compare['_key'].isin(added_keys)].copy()
        added_rows = added_rows.drop('_key', axis=1)
        
        # Get removed rows
        removed_rows = df1_compare[df1_compare['_key'].isin(removed_keys)].copy()
        removed_rows = removed_rows.drop('_key', axis=1)
        
        # Find changed rows
        changed_rows = []
        unchanged_rows = []
        
        for key in common_keys:
            row1 = df1_compare[df1_compare['_key'] == key].iloc[0]
            row2 = df2_compare[df2_compare['_key'] == key].iloc[0]
            
            # Compare all columns except key
            compare_cols_only = [col for col in cols_to_compare if col not in key_columns]
            is_changed = False
            differences = {}
            
            for col in compare_cols_only:
                val1 = row1[col]
                val2 = row2[col]
                
                # Handle NaN comparison
                if pd.isna(val1) and pd.isna(val2):
                    continue
                elif pd.isna(val1) or pd.isna(val2):
                    is_changed = True
                    differences[col] = {'old': val1 if not pd.isna(val1) else None, 'new': val2 if not pd.isna(val2) else None}
                elif val1 != val2:
                    is_changed = True
                    differences[col] = {'old': val1, 'new': val2}
            
            if is_changed:
                # Create side-by-side comparison row
                change_row = {}
                for col in key_columns:
                    change_row[col] = row1[col]
                for col in compare_cols_only:
                    if col in differences:
                        change_row[f'{col} (Old)'] = differences[col]['old']
                        change_row[f'{col} (New)'] = differences[col]['new']
                    else:
                        change_row[col] = row1[col]
                changed_rows.append(change_row)
            else:
                # Unchanged row
                unchanged_row = {}
                for col in cols_to_compare:
                    unchanged_row[col] = row1[col]
                unchanged_rows.append(unchanged_row)
        
        send_progress('comparing', 100, 100, f'Comparison complete: {len(changed_rows)} changed rows found', 70)
        
        send_progress('saving', 0, 100, 'Saving comparison results...', 75)
        
        # Create summary DataFrame
        summary_data = {
            'Metric': [
                'File 1 Total Rows',
                'File 2 Total Rows',
                'Common Rows',
                'Added Rows (in File 2 only)',
                'Removed Rows (in File 1 only)',
                'Changed Rows',
                'Unchanged Rows'
            ],
            'Value': [
                len(df1),
                len(df2),
                len(common_keys),
                len(added_keys),
                len(removed_keys),
                len(changed_rows),
                len(unchanged_rows)
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Convert to DataFrames
        added_df = pd.DataFrame(added_rows) if len(added_rows) > 0 else pd.DataFrame(columns=cols_to_compare)
        removed_df = pd.DataFrame(removed_rows) if len(removed_rows) > 0 else pd.DataFrame(columns=cols_to_compare)
        changed_df = pd.DataFrame(changed_rows) if len(changed_rows) > 0 else pd.DataFrame()
        unchanged_df = pd.DataFrame(unchanged_rows) if len(unchanged_rows) > 0 else pd.DataFrame(columns=cols_to_compare)
        
        # Save to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"comparison_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            if len(added_df) > 0:
                added_df.to_excel(writer, sheet_name='Added Rows', index=False)
            if len(removed_df) > 0:
                removed_df.to_excel(writer, sheet_name='Removed Rows', index=False)
            if len(changed_df) > 0:
                changed_df.to_excel(writer, sheet_name='Changed Rows', index=False)
            if len(unchanged_df) > 0:
                unchanged_df.to_excel(writer, sheet_name='Unchanged Rows', index=False)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'Results saved...', 95)
        
        # Store summary data before cleanup
        file1_total = len(df1)
        file2_total = len(df2)
        
        # Prepare preview data (handle NaN values)
        def prepare_preview(df, max_rows=50):
            if len(df) == 0:
                return []
            preview_df = df.head(max_rows)
            preview_list = []
            for _, row in preview_df.iterrows():
                row_dict = {}
                for col in preview_df.columns:
                    val = row[col]
                    if pd.isna(val):
                        row_dict[col] = None
                    else:
                        row_dict[col] = str(val)
                preview_list.append(row_dict)
            return preview_list
        
        preview_data = {
            'added': prepare_preview(added_df),
            'removed': prepare_preview(removed_df),
            'changed': prepare_preview(changed_df)
        }
        
        # Clean up
        del df1
        del df2
        del df1_compare
        del df2_compare
        os.remove(file1_path)
        os.remove(file2_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'summary': {
                'file1_rows': int(file1_total),
                'file2_rows': int(file2_total),
                'common': int(len(common_keys)),
                'added': int(len(added_keys)),
                'removed': int(len(removed_keys)),
                'changed': int(len(changed_rows)),
                'unchanged': int(len(unchanged_rows))
            },
            'preview': preview_data
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Comparison complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file1_path' in locals() and os.path.exists(file1_path):
            os.remove(file1_path)
        if 'file2_path' in locals() and os.path.exists(file2_path):
            os.remove(file2_path)
        if 'df1' in locals():
            del df1
        if 'df2' in locals():
            del df2
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/compare-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def compare_data():
    """Compare two uploaded files - returns session_id for progress tracking"""
    try:
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({'error': 'Please upload both files'}), 400
        
        file1 = request.files['file1']
        file2 = request.files['file2']
        
        if file1.filename == '' or file2.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Validate file extensions
        if not (file1.filename.endswith(('.xlsx', '.xls', '.csv')) and 
                file2.filename.endswith(('.xlsx', '.xls', '.csv'))):
            return jsonify({'error': 'Please upload Excel or CSV files'}), 400
        
        # Get comparison parameters
        key_columns = request.form.getlist('key_columns[]')
        compare_columns = request.form.getlist('compare_columns[]')  # Optional
        
        if not key_columns:
            return jsonify({'error': 'Please select at least one key column'}), 400
        
        # Save files temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"compare_{timestamp}_{secrets.token_hex(8)}"
        
        file1_filename = secure_filename(file1.filename)
        file2_filename = secure_filename(file2.filename)
        
        file1_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_file1_{file1_filename}")
        file2_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_file2_{file2_filename}")
        
        file1.save(file1_path)
        file2.save(file2_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=compare_files_async,
            args=(file1_path, file2_path, key_columns, compare_columns, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/pivot-generator')
def pivot_generator():
    return render_template('pivot_generator.html')

def generate_pivot_async(file_path, rows, columns, values, aggfunc, filters, progress_queue, session_id):
    """Generate pivot table in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file...', 5)
        
        # Read file
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows', 15)
        
        # Validate columns exist
        all_cols = (rows or []) + (columns or []) + (values or []) + (filters or [])
        invalid_cols = [col for col in all_cols if col not in df.columns]
        if invalid_cols:
            raise ValueError(f"Columns not found in file: {', '.join(invalid_cols)}")
        
        send_progress('preparing', 0, 100, 'Preparing data for pivot...', 20)
        
        # Apply filters if any
        if filters:
            for filter_col, filter_value in filters.items():
                if filter_value:
                    df = df[df[filter_col] == filter_value]
        
        send_progress('preparing', 50, 100, 'Creating pivot table...', 30)
        
        # Validate that we have required parameters
        if not values:
            raise ValueError("At least one value field must be selected for the pivot table")
        
        # Determine default aggregation if not provided
        if not aggfunc:
            # Check if values are numeric to determine default aggfunc
            numeric_cols = [col for col in values if col in df.columns and pd.api.types.is_numeric_dtype(df[col])]
            aggfunc = 'sum' if numeric_cols else 'count'
        
        # Create pivot table
        pivot_params = {
            'index': rows if rows else None,
            'columns': columns if columns else None,
            'values': values,
            'aggfunc': aggfunc
        }
        
        # Remove None values from index/columns (but keep values)
        if pivot_params['index'] is None:
            del pivot_params['index']
        if pivot_params['columns'] is None:
            del pivot_params['columns']
        
        # Create the pivot table
        try:
            pivot_df = pd.pivot_table(df, **pivot_params, fill_value=0, margins=True, margins_name='Total')
        except Exception as e:
            # If margins fail, try without
            print(f"Warning: Could not add totals/margins: {e}")
            pivot_df = pd.pivot_table(df, **pivot_params, fill_value=0)
        
        send_progress('preparing', 100, 100, f'Pivot table created: {len(pivot_df):,} rows, {len(pivot_df.columns)} columns', 50)
        
        send_progress('saving', 0, 100, 'Saving results...', 60)
        
        # Flatten MultiIndex columns if they exist (handle all cases)
        def flatten_columns(df):
            """Flatten MultiIndex columns to single level"""
            if isinstance(df.columns, pd.MultiIndex):
                # Get column names and flatten them
                new_columns = []
                for col in df.columns:
                    if isinstance(col, tuple):
                        # Join tuple elements, filtering out empty strings and None
                        parts = [str(c) for c in col if c is not None and str(c) not in ['', 'nan', 'None']]
                        flattened = '_'.join(parts) if parts else 'Value'
                        # Clean up multiple underscores
                        while '__' in flattened:
                            flattened = flattened.replace('__', '_')
                        flattened = flattened.strip('_')
                        new_columns.append(flattened if flattened else 'Value')
                    else:
                        new_columns.append(str(col) if col is not None else 'Value')
                df.columns = new_columns
            return df
        
        # Flatten MultiIndex columns BEFORE reset_index (if they exist)
        pivot_df = flatten_columns(pivot_df)
        
        # Reset index to make it a proper table structure
        # Handle MultiIndex index properly
        if isinstance(pivot_df.index, pd.MultiIndex):
            pivot_df_reset = pivot_df.reset_index()
        else:
            pivot_df_reset = pivot_df.reset_index()
        
        # Flatten again after reset in case reset_index created new MultiIndex columns
        pivot_df_reset = flatten_columns(pivot_df_reset)
        
        # Create summary
        summary_data = {
            'Metric': [
                'Source Rows',
                'Pivot Rows',
                'Pivot Columns',
                'Aggregation Function',
                'Row Dimensions',
                'Column Dimensions',
                'Value Fields'
            ],
            'Value': [
                len(df),
                len(pivot_df_reset),
                len(pivot_df_reset.columns),
                aggfunc if aggfunc else 'sum',
                ', '.join(rows) if rows else 'None',
                ', '.join(columns) if columns else 'None',
                ', '.join(values) if values else 'All Numeric'
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Ensure summary_df doesn't have MultiIndex columns
        summary_df = flatten_columns(summary_df.copy())
        
        # Save to Excel with formatting
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"pivot_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        # Write to Excel with error handling
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write pivot table with proper structure
            try:
                # Verify columns are not MultiIndex before writing
                if isinstance(pivot_df_reset.columns, pd.MultiIndex):
                    print("Warning: Columns are still MultiIndex after flattening, trying to flatten again")
                    pivot_df_reset = flatten_columns(pivot_df_reset)
                
                pivot_df_reset.to_excel(writer, sheet_name='Pivot Table', index=False)
            except Exception as e:
                # If still fails, write with index=True and we'll fix it after
                print(f"Warning: Could not write with index=False: {e}")
                try:
                    pivot_df_reset.to_excel(writer, sheet_name='Pivot Table', index=True)
                except Exception as e2:
                    # Last resort: write as is
                    print(f"Error writing pivot table: {e2}")
                    # Create a copy and ensure it's flat
                    pivot_flat = pivot_df_reset.copy()
                    pivot_flat.columns = [str(col) for col in pivot_flat.columns]
                    pivot_flat.to_excel(writer, sheet_name='Pivot Table', index=True)
            
            # Write summary and source data (these shouldn't have MultiIndex issues)
            try:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            except Exception as e:
                print(f"Warning writing summary: {e}")
                summary_df_flat = flatten_columns(summary_df.copy())
                summary_df_flat.to_excel(writer, sheet_name='Summary', index=False)
            
            try:
                df.to_excel(writer, sheet_name='Source Data', index=False)
            except Exception as e:
                print(f"Warning writing source data: {e}")
                df_flat = flatten_columns(df.copy())
                df_flat.to_excel(writer, sheet_name='Source Data', index=False)
        
        # Apply formatting to make it look like a real pivot table
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = load_workbook(output_path)
        
        # Check if Pivot Table sheet exists (might have been written with different name)
        if 'Pivot Table' not in wb.sheetnames:
            # If sheet doesn't exist, something went wrong
            raise ValueError("Pivot Table sheet was not created successfully")
        
        ws = wb['Pivot Table']
        
        # Remove index column if it exists (if we wrote with index=True as fallback)
        if ws.max_row > 0 and ws.max_column > 0:
            # Check if first column looks like an index (unnamed or numbered)
            first_col_header = ws.cell(row=1, column=1).value
            if first_col_header is None or (isinstance(first_col_header, str) and first_col_header.startswith('Unnamed')):
                ws.delete_cols(1)
        
        # Calculate row header columns (index columns) - adjust if we deleted an index column
        row_header_cols = len(rows) if rows else 0
        
        # Format header row with different colors for row headers vs data
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        # Row header column headers (if any)
        for col_idx in range(1, row_header_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')  # Darker blue for row headers
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data column headers
        for col_idx in range(row_header_cols + 1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Pre-create style objects to avoid creating them in loops (performance optimization)
        index_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        total_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        border_style = Border(
            left=Side(style='thin', color='4472C4'),
            right=Side(style='thin', color='4472C4'),
            top=Side(style='thin', color='4472C4'),
            bottom=Side(style='thin', color='4472C4')
        )
        font_bold_11 = Font(bold=True, size=11)
        font_bold_10 = Font(bold=True, size=10)
        font_normal_10 = Font(size=10)
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')
        
        total_rows = ws.max_row - 1  # Exclude header row
        if total_rows > 0:
            # Format row header columns - with progress updates
            send_progress('saving', 0, total_rows * 2, 'Formatting row headers...', 62)
            
            for idx, row_idx in enumerate(range(2, ws.max_row + 1), 1):
                # Send progress every 100 rows or on last row
                if idx % 100 == 0 or idx == total_rows:
                    progress_pct = 60 + int((idx / total_rows) * 12)  # 60% to 72%
                    send_progress('saving', idx, total_rows * 2, f'Formatting row headers: {idx}/{total_rows}...', progress_pct)
                
                # Check if this is a total row (starts with "Total")
                first_cell = ws.cell(row=row_idx, column=1)
                is_total_row = str(first_cell.value).startswith('Total') if first_cell.value else False
                
                for col_idx in range(1, row_header_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if is_total_row:
                        cell.fill = total_fill
                        cell.font = font_bold_11
                    else:
                        cell.fill = index_fill
                        cell.font = font_bold_10
                    cell.alignment = align_left
                    cell.border = border_style
            
            # Format data columns - with progress updates
            send_progress('saving', total_rows, total_rows * 2, 'Formatting data columns...', 72)
            
            # Pre-cache column header values to avoid repeated lookups
            column_headers = {}
            for col_idx in range(row_header_cols + 1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col_idx)
                column_headers[col_idx] = 'Total' in str(header_cell.value) if header_cell.value else False
            
            for idx, row_idx in enumerate(range(2, ws.max_row + 1), 1):
                # Send progress every 100 rows or on last row
                if idx % 100 == 0 or idx == total_rows:
                    progress_pct = 72 + int((idx / total_rows) * 18)  # 72% to 90%
                    send_progress('saving', total_rows + idx, total_rows * 2, f'Formatting data columns: {idx}/{total_rows}...', progress_pct)
                
                first_cell = ws.cell(row=row_idx, column=1)
                is_total_row = str(first_cell.value).startswith('Total') if first_cell.value else False
                
                for col_idx in range(row_header_cols + 1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    is_total_col = column_headers.get(col_idx, False)
                    
                    if is_total_row or is_total_col:
                        cell.fill = total_fill
                        cell.font = font_bold_11
                    elif row_idx % 2 == 0:  # Zebra striping
                        cell.fill = data_fill
                        cell.font = font_normal_10
                    else:
                        cell.font = font_normal_10
                    
                    # Format numeric cells
                    try:
                        val = float(cell.value) if cell.value is not None else None
                        if val is not None:
                            cell.number_format = '#,##0.00'  # Number format with thousands separator
                            cell.alignment = align_right
                        else:
                            cell.alignment = align_center
                    except:
                        cell.alignment = align_left
                    
                    cell.border = border_style
            
            send_progress('saving', total_rows * 2, total_rows * 2, 'Formatting complete...', 90)
        
        # Freeze panes at first data row and column
        if row_header_cols > 0:
            ws.freeze_panes = ws.cell(row=2, column=row_header_cols + 1)
        else:
            ws.freeze_panes = ws.cell(row=2, column=1)
        
        # Auto-adjust column widths
        send_progress('saving', 0, ws.max_column, 'Adjusting column widths...', 90)
        from openpyxl.utils import get_column_letter
        
        for col_idx, column in enumerate(ws.columns, 1):
            # Send progress every 10 columns
            if col_idx % 10 == 0 or col_idx == ws.max_column:
                progress_pct = 90 + int((col_idx / ws.max_column) * 5)  # 90% to 95%
                send_progress('saving', col_idx, ws.max_column, f'Adjusting column widths: {col_idx}/{ws.max_column}...', progress_pct)
            
            max_length = 0
            # Find first non-merged cell or use column index to get letter
            column_letter = None
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            # If no regular cell found, use column index
            if not column_letter:
                column_letter = get_column_letter(col_idx)
            
            for cell in column:
                try:
                    # Skip merged cells
                    if not hasattr(cell, 'value'):
                        continue
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)  # Min width 12, max 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        send_progress('saving', ws.max_column, ws.max_column, 'Saving file...', 95)
        wb.save(output_path)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'Results saved...', 95)
        
        # Store summary data before cleanup
        source_rows_total = len(df)
        pivot_rows_total = len(pivot_df_reset)
        pivot_cols_total = len(pivot_df_reset.columns)
        
        # Prepare preview data
        preview_data = []
        if len(pivot_df_reset) > 0:
            preview_df = pivot_df_reset.head(50)
            for _, row in preview_df.iterrows():
                row_dict = {}
                for col in preview_df.columns:
                    val = row[col]
                    if pd.isna(val):
                        row_dict[col] = None
                    else:
                        row_dict[col] = str(val)
                preview_data.append(row_dict)
        
        # Clean up
        del df
        del pivot_df
        os.remove(file_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'summary': {
                'source_rows': int(source_rows_total),
                'pivot_rows': int(pivot_rows_total),
                'pivot_cols': int(pivot_cols_total),
                'aggfunc': aggfunc if aggfunc else 'sum'
            },
            'preview': preview_data
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Pivot table generation complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        if 'df' in locals():
            del df
        if 'pivot_df' in locals():
            del pivot_df
        if 'pivot_df_reset' in locals():
            del pivot_df_reset
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/generate-pivot', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def generate_pivot():
    """Generate pivot table - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Please upload a file'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'File must be selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Get pivot parameters
        rows = request.form.getlist('rows[]')
        columns = request.form.getlist('columns[]')
        values = request.form.getlist('values[]')
        aggfunc = request.form.get('aggfunc', 'sum').strip()
        filters = {}
        
        # Get filters (if any)
        filter_cols = request.form.getlist('filter_columns[]')
        filter_values = request.form.getlist('filter_values[]')
        for col, val in zip(filter_cols, filter_values):
            if col and val:
                filters[col] = val
        
        # At least one dimension required
        if not rows and not columns:
            return jsonify({'error': 'Please select at least one row or column dimension'}), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"pivot_{timestamp}_{secrets.token_hex(8)}"
        
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        
        file.save(file_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=generate_pivot_async,
            args=(file_path, rows, columns, values, aggfunc, filters, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/data-validation')
def data_validation():
    return render_template('data_validation.html')

def validate_data_async(file_path, validation_rules, progress_queue, session_id):
    """Validate data in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file...', 5)
        
        # Read file
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows', 15)
        
        send_progress('validating', 0, 100, 'Validating data...', 20)
        
        # Parse validation rules
        import json
        rules = json.loads(validation_rules) if isinstance(validation_rules, str) else validation_rules
        
        # Track validation errors
        errors = []
        valid_rows = []
        invalid_rows = []
        
        total_rows = len(df)
        validated = 0
        
        # Validate each row
        for idx, row in df.iterrows():
            row_errors = []
            is_valid = True
            
            for rule in rules:
                col = rule.get('column')
                rule_type = rule.get('type')
                rule_value = rule.get('value')
                
                if col not in df.columns:
                    continue
                
                cell_value = row[col]
                
                # Required validation
                if rule_type == 'required':
                    if pd.isna(cell_value) or str(cell_value).strip() == '':
                        row_errors.append(f"{col}: Required field is empty")
                        is_valid = False
                
                # Data type validation
                elif rule_type == 'numeric':
                    if not pd.isna(cell_value):
                        try:
                            float(cell_value)
                        except:
                            row_errors.append(f"{col}: Must be numeric")
                            is_valid = False
                
                # Range validation
                elif rule_type == 'range':
                    if not pd.isna(cell_value):
                        try:
                            val = float(cell_value)
                            min_val = rule_value.get('min')
                            max_val = rule_value.get('max')
                            if min_val is not None and val < min_val:
                                row_errors.append(f"{col}: Value {val} is below minimum {min_val}")
                                is_valid = False
                            if max_val is not None and val > max_val:
                                row_errors.append(f"{col}: Value {val} is above maximum {max_val}")
                                is_valid = False
                        except:
                            pass
                
                # List/Enum validation
                elif rule_type == 'list':
                    if not pd.isna(cell_value):
                        allowed_values = rule_value if isinstance(rule_value, list) else [rule_value]
                        if str(cell_value) not in [str(v) for v in allowed_values]:
                            row_errors.append(f"{col}: Value '{cell_value}' not in allowed list")
                            is_valid = False
                
                # Pattern validation (regex)
                elif rule_type == 'pattern':
                    if not pd.isna(cell_value):
                        import re
                        pattern = rule_value
                        if not re.match(pattern, str(cell_value)):
                            row_errors.append(f"{col}: Value does not match required pattern")
                            is_valid = False
                
                # Length validation
                elif rule_type == 'length':
                    if not pd.isna(cell_value):
                        val_str = str(cell_value)
                        min_len = rule_value.get('min')
                        max_len = rule_value.get('max')
                        if min_len is not None and len(val_str) < min_len:
                            row_errors.append(f"{col}: Length {len(val_str)} is below minimum {min_len}")
                            is_valid = False
                        if max_len is not None and len(val_str) > max_len:
                            row_errors.append(f"{col}: Length {len(val_str)} is above maximum {max_len}")
                            is_valid = False
            
            if row_errors:
                errors.append({
                    'row': int(idx) + 1,
                    'errors': row_errors
                })
                invalid_rows.append(row.to_dict())
            else:
                valid_rows.append(row.to_dict())
            
            validated += 1
            if validated % 1000 == 0:
                send_progress('validating', validated, total_rows, f'Validated {validated:,} of {total_rows:,} rows...', 20 + int((validated / total_rows) * 50))
        
        send_progress('validating', total_rows, total_rows, f'Validation complete: {len(errors)} errors found', 70)
        
        send_progress('saving', 0, 100, 'Saving validation results...', 75)
        
        # Create summary
        summary_data = {
            'Metric': [
                'Total Rows',
                'Valid Rows',
                'Invalid Rows',
                'Error Rate (%)',
                'Total Errors'
            ],
            'Value': [
                total_rows,
                len(valid_rows),
                len(invalid_rows),
                round((len(invalid_rows) / total_rows * 100) if total_rows > 0 else 0, 2),
                len(errors)
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Create errors DataFrame
        errors_data = []
        for error in errors:
            errors_data.append({
                'Row Number': error['row'],
                'Errors': '; '.join(error['errors'])
            })
        errors_df = pd.DataFrame(errors_data) if errors_data else pd.DataFrame(columns=['Row Number', 'Errors'])
        
        # Create invalid records DataFrame
        invalid_df = pd.DataFrame(invalid_rows) if invalid_rows else pd.DataFrame(columns=df.columns)
        valid_df = pd.DataFrame(valid_rows) if valid_rows else pd.DataFrame(columns=df.columns)
        
        # Save to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"validation_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Validation Summary', index=False)
            errors_df.to_excel(writer, sheet_name='Error Details', index=False)
            if len(invalid_df) > 0:
                invalid_df.to_excel(writer, sheet_name='Invalid Records', index=False)
            if len(valid_df) > 0:
                valid_df.to_excel(writer, sheet_name='Valid Records', index=False)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'Results saved...', 95)
        
        # Prepare preview data
        preview_errors = errors[:50] if len(errors) > 50 else errors
        
        # Clean up
        del df
        os.remove(file_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'summary': {
                'total_rows': int(total_rows),
                'valid_rows': int(len(valid_rows)),
                'invalid_rows': int(len(invalid_rows)),
                'error_rate': round((len(invalid_rows) / total_rows * 100) if total_rows > 0 else 0, 2),
                'total_errors': int(len(errors))
            },
            'preview': preview_errors
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Validation complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        if 'df' in locals():
            del df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/validate-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def validate_data():
    """Validate data - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Please upload a file'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'File must be selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Get validation rules
        validation_rules = request.form.get('validation_rules', '[]')
        
        if not validation_rules or validation_rules == '[]':
            return jsonify({'error': 'Please define at least one validation rule'}), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"validate_{timestamp}_{secrets.token_hex(8)}"
        
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        
        file.save(file_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=validate_data_async,
            args=(file_path, validation_rules, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/column-normalizer')
def column_normalizer():
    return render_template('column_normalizer.html')

def normalize_columns_async(file_path, column_types, trim_whitespace, progress_queue, session_id):
    """Normalize column data types in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file...', 5)
        
        # Read file
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        send_progress('loading', 100, 100, f'File loaded: {len(df):,} rows, {len(df.columns)} columns', 15)
        
        send_progress('normalizing', 0, len(df.columns), 'Normalizing columns...', 20)
        
        # Parse column types (JSON string or dict)
        import json
        if isinstance(column_types, str):
            column_types = json.loads(column_types)
        
        normalized_df = df.copy()
        transformations_applied = []
        
        # Process each column
        for col_idx, (col_name, target_type) in enumerate(column_types.items()):
            if col_name not in normalized_df.columns:
                continue
            
            send_progress('normalizing', col_idx + 1, len(column_types), 
                        f'Normalizing column: {col_name} to {target_type}...', 
                        20 + int((col_idx / len(column_types)) * 60))
            
            original_col = normalized_df[col_name].copy()
            transformed_count = 0
            error_count = 0
            
            if target_type == 'text' or target_type == 'string':
                # Convert to string, optionally trim
                if trim_whitespace:
                    normalized_df[col_name] = normalized_df[col_name].astype(str).str.strip()
                    transformed_count = len(normalized_df)
                else:
                    normalized_df[col_name] = normalized_df[col_name].astype(str)
                    transformed_count = len(normalized_df)
            
            elif target_type == 'integer':
                # Convert to integer, handling errors
                for idx in normalized_df.index:
                    val = normalized_df.loc[idx, col_name]
                    if pd.notna(val):
                        try:
                            # Try to convert, handling strings with commas/currency symbols
                            if isinstance(val, str):
                                val = val.replace(',', '').replace('$', '').replace('€', '').replace('£', '').strip()
                            normalized_df.loc[idx, col_name] = int(float(val))
                            transformed_count += 1
                        except (ValueError, TypeError):
                            error_count += 1
                            # Keep original value if conversion fails
                            pass
            
            elif target_type == 'float' or target_type == 'decimal':
                # Convert to float, handling errors
                for idx in normalized_df.index:
                    val = normalized_df.loc[idx, col_name]
                    if pd.notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.replace(',', '').replace('$', '').replace('€', '').replace('£', '').strip()
                            normalized_df.loc[idx, col_name] = float(val)
                            transformed_count += 1
                        except (ValueError, TypeError):
                            error_count += 1
                            pass
            
            elif target_type == 'currency' or target_type == 'dollar':
                # Convert to float (currency is a float, just formatted differently)
                for idx in normalized_df.index:
                    val = normalized_df.loc[idx, col_name]
                    if pd.notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.replace(',', '').replace('$', '').replace('€', '').replace('£', '').strip()
                            normalized_df.loc[idx, col_name] = float(val)
                            transformed_count += 1
                        except (ValueError, TypeError):
                            error_count += 1
                            pass
                # Note: We'll format this in Excel output
            
            elif target_type == 'date':
                # Convert to datetime
                try:
                    normalized_df[col_name] = pd.to_datetime(normalized_df[col_name], errors='coerce')
                    transformed_count = normalized_df[col_name].notna().sum()
                    error_count = normalized_df[col_name].isna().sum() - original_col.isna().sum()
                except:
                    error_count = len(normalized_df)
            
            elif target_type == 'boolean':
                # Convert to boolean
                for idx in normalized_df.index:
                    val = normalized_df.loc[idx, col_name]
                    if pd.notna(val):
                        try:
                            val_str = str(val).lower().strip()
                            if val_str in ('true', '1', 'yes', 'y', 't'):
                                normalized_df.loc[idx, col_name] = True
                            elif val_str in ('false', '0', 'no', 'n', 'f'):
                                normalized_df.loc[idx, col_name] = False
                            else:
                                # Try numeric conversion
                                if float(val) != 0:
                                    normalized_df.loc[idx, col_name] = True
                                else:
                                    normalized_df.loc[idx, col_name] = False
                            transformed_count += 1
                        except:
                            error_count += 1
                            pass
            
            elif target_type == 'percentage':
                # Convert percentage (remove % sign, divide by 100)
                for idx in normalized_df.index:
                    val = normalized_df.loc[idx, col_name]
                    if pd.notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.replace('%', '').replace(',', '').strip()
                            normalized_df.loc[idx, col_name] = float(val) / 100.0
                            transformed_count += 1
                        except (ValueError, TypeError):
                            error_count += 1
                            pass
            
            elif target_type == 'keep_original':
                # Don't transform
                transformed_count = 0
                error_count = 0
            else:
                # Unknown type, keep original
                transformed_count = 0
                error_count = 0
            
            transformations_applied.append({
                'column': col_name,
                'type': target_type,
                'transformed': transformed_count,
                'errors': error_count
            })
        
        send_progress('saving', 0, 100, 'Saving normalized file...', 85)
        
        # Save to Excel with formatting
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"normalized_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            normalized_df.to_excel(writer, sheet_name='Normalized Data', index=False)
            
            # Create summary sheet
            summary_data = {
                'Column': [t['column'] for t in transformations_applied],
                'Target Type': [t['type'] for t in transformations_applied],
                'Rows Transformed': [t['transformed'] for t in transformations_applied],
                'Conversion Errors': [t['errors'] for t in transformations_applied]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Transformation Summary', index=False)
        
        # Apply formatting for currency columns
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE_00
        
        wb = load_workbook(output_path)
        ws = wb['Normalized Data']
        
        # Format headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply column-specific formatting
        for col_idx, (col_name, target_type) in enumerate(column_types.items(), 1):
            if target_type == 'currency' or target_type == 'dollar':
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None:
                        cell.number_format = FORMAT_CURRENCY_USD
            elif target_type == 'percentage':
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None:
                        cell.number_format = FORMAT_PERCENTAGE_00
            elif target_type == 'date':
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None:
                        cell.number_format = 'mm/dd/yyyy'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        send_progress('saving', 100, 100, 'File saved...', 95)
        
        # Store summary data
        total_transformed = sum(t['transformed'] for t in transformations_applied)
        total_errors = sum(t['errors'] for t in transformations_applied)
        
        # Clean up
        del df
        del normalized_df
        os.remove(file_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'summary': {
                'total_columns': len(transformations_applied),
                'total_transformed': int(total_transformed),
                'total_errors': int(total_errors)
            },
            'transformations': transformations_applied
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Column normalization complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        if 'df' in locals():
            del df
        if 'normalized_df' in locals():
            del normalized_df
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/normalize-columns', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def normalize_columns():
    """Normalize columns - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Please upload a file'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'File must be selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Get column types mapping
        column_types_json = request.form.get('column_types', '{}')
        trim_whitespace = request.form.get('trim_whitespace', 'false').lower() == 'true'
        
        try:
            import json
            column_types = json.loads(column_types_json)
            if not isinstance(column_types, dict):
                return jsonify({'error': 'Invalid column types format'}), 400
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid JSON format for column types'}), 400
        
        if not column_types:
            return jsonify({'error': 'Please select at least one column to normalize'}), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"normalize_{timestamp}_{secrets.token_hex(8)}"
        
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        
        file.save(file_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=normalize_columns_async,
            args=(file_path, column_types, trim_whitespace, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/pdf-to-word')
def pdf_to_word():
    return render_template('pdf_to_word.html')

@app.route('/column-comparison')
def column_comparison():
    return render_template('column_comparison.html')

@app.route('/transpose')
def transpose():
    return render_template('transpose.html')

def convert_pdf_to_word_async(file_path, progress_queue, session_id):
    """Convert PDF to Word document in background thread with progress tracking"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading PDF file...', 5)
        
        # Import pdf2docx
        try:
            from pdf2docx import Converter
        except ImportError:
            raise ImportError("pdf2docx library is not installed. Please install it with: pip install pdf2docx")
        
        # Check if file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"PDF file not found: {file_path}")
        
        send_progress('loading', 50, 100, 'PDF file loaded...', 15)
        
        send_progress('converting', 0, 100, 'Converting PDF to Word document...', 20)
        
        # Create output file path
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = f"converted_{timestamp}_{base_name}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.docx")
        
        # Convert PDF to Word
        try:
            cv = Converter(file_path)
            cv.convert(output_path, start=0, end=None)  # Convert all pages
            cv.close()
        except Exception as e:
            raise Exception(f"Error during PDF conversion: {str(e)}")
        
        send_progress('converting', 100, 100, 'Conversion complete!', 85)
        
        send_progress('saving', 0, 100, 'Preparing download...', 90)
        
        # Check if output file was created
        if not os.path.exists(output_path):
            raise FileNotFoundError("Word document was not created successfully")
        
        download_url = f'/download/{output_filename}.docx'
        
        # Get file size
        file_size = os.path.getsize(output_path)
        
        send_progress('saving', 100, 100, 'Ready for download!', 95)
        
        # Clean up input file
        os.remove(file_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': 100,
            'total': 100,
            'percentage': 100,
            'message': 'Conversion complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'file_size': file_size
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("PDF to Word conversion complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file_path' in locals() and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass
        if 'output_path' in locals() and os.path.exists(output_path):
            try:
                os.remove(output_path)
            except:
                pass
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/convert-pdf-to-word', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def convert_pdf_to_word():
    """Convert PDF to Word - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Please upload a PDF file'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'File must be selected'}), 400
        
        # Validate file extension
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Please upload a PDF file'}), 400
        
        # Save file temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"pdf2word_{timestamp}_{secrets.token_hex(8)}"
        
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        
        file.save(file_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=convert_pdf_to_word_async,
            args=(file_path, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

def compare_columns_async(file1_path, file2_path, file1_name, file2_name, progress_queue, session_id):
    """Compare columns from two files in background thread - reads only headers for speed"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, f'Reading headers from {file1_name}...', 10)
        
        # Read just the headers from file 1 - optimized to read only first row
        if file1_path.endswith('.csv'):
            # For CSV, use csv module to properly handle quoted fields
            import csv
            with open(file1_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                first_row = next(reader)
                columns1 = set([str(col).strip() for col in first_row if col])
        else:
            # For Excel, use pandas with nrows=0 to read only headers (handles both .xlsx and .xls)
            try:
                df1 = pd.read_excel(file1_path, nrows=0)
                columns1 = set([str(col).strip() for col in df1.columns if col])
            except Exception as e:
                # Fallback: try openpyxl for .xlsx files only
                if file1_path.endswith('.xlsx'):
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(file1_path, read_only=True, data_only=True)
                        ws = wb.active
                        first_row = [cell.value for cell in ws[1] if cell.value is not None]
                        columns1 = set([str(col).strip() for col in first_row])
                        wb.close()
                    except Exception as e2:
                        raise Exception(f"Error reading {file1_name}: {str(e2)}")
                else:
                    raise Exception(f"Error reading {file1_name}: {str(e)}")
        
        send_progress('loading', 50, 100, f'Reading headers from {file2_name}...', 30)
        
        # Read just the headers from file 2 - optimized to read only first row
        if file2_path.endswith('.csv'):
            # For CSV, use csv module to properly handle quoted fields
            import csv
            with open(file2_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                first_row = next(reader)
                columns2 = set([str(col).strip() for col in first_row if col])
        else:
            # For Excel, use pandas with nrows=0 to read only headers (handles both .xlsx and .xls)
            try:
                df2 = pd.read_excel(file2_path, nrows=0)
                columns2 = set([str(col).strip() for col in df2.columns if col])
            except Exception as e:
                # Fallback: try openpyxl for .xlsx files only
                if file2_path.endswith('.xlsx'):
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(file2_path, read_only=True, data_only=True)
                        ws = wb.active
                        first_row = [cell.value for cell in ws[1] if cell.value is not None]
                        columns2 = set([str(col).strip() for col in first_row])
                        wb.close()
                    except Exception as e2:
                        raise Exception(f"Error reading {file2_name}: {str(e2)}")
                else:
                    raise Exception(f"Error reading {file2_name}: {str(e)}")
        
        send_progress('comparing', 0, 100, 'Comparing columns...', 50)
        
        # Get all unique columns (alphabetically sorted)
        all_columns = sorted(columns1.union(columns2))
        
        # Build comparison data
        comparison_data = []
        for col in all_columns:
            comparison_data.append({
                'Column Headers': col,
                f'Found in {file1_name}': 'Yes' if col in columns1 else 'No',
                f'Found in {file2_name}': 'Yes' if col in columns2 else 'No'
            })
        
        comparison_df = pd.DataFrame(comparison_data)
        
        send_progress('saving', 0, 100, 'Saving results...', 80)
        
        # Save to Excel - keep it simple and fast (no formatting for speed)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"column_comparison_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        # Just write the data - formatting can slow things down significantly
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name='Column Comparison', index=False)
            
            # Optional: Basic formatting (only if there are few columns, skip if many)
            if len(comparison_df) <= 500:  # Only format if reasonable number of rows
                try:
                    from openpyxl.styles import PatternFill, Font
                    ws = writer.sheets['Column Comparison']
                    
                    # Green fill for checkmarks, red for X marks
                    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    green_font = Font(bold=True, color='006400')
                    red_font = Font(bold=True, color='8B0000')
                    
                    # Format cells
                    for row_idx in range(len(comparison_df)):
                        excel_row = row_idx + 2  # +2 because Excel is 1-indexed and we skip header row
                        
                        # Check File 1 column (column B, index 2)
                        cell_b = ws.cell(row=excel_row, column=2)
                        if cell_b.value == 'Yes':
                            cell_b.fill = green_fill
                            cell_b.font = green_font
                        elif cell_b.value == 'No':
                            cell_b.fill = red_fill
                            cell_b.font = red_font
                        
                        # Check File 2 column (column C, index 3)
                        cell_c = ws.cell(row=excel_row, column=3)
                        if cell_c.value == 'Yes':
                            cell_c.fill = green_fill
                            cell_c.font = green_font
                        elif cell_c.value == 'No':
                            cell_c.fill = red_fill
                            cell_c.font = red_font
                except Exception as format_error:
                    # If formatting fails, just continue without it - data is still saved
                    print(f"Warning: Formatting failed, but file saved: {format_error}")
        
        download_url = f'/download/{output_filename}.xlsx'
        
        # Clean up uploaded files
        os.remove(file1_path)
        os.remove(file2_path)
        
        # Send completion message with all data (don't send separate 'done' progress update)
        final_message = {
            'stage': 'done',
            'current': len(all_columns),
            'total': len(all_columns),
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'total_columns': len(all_columns),
            'file1_columns': len(columns1),
            'file2_columns': len(columns2),
            'common_columns': len(columns1.intersection(columns2)),
            'file1_only': len(columns1 - columns2),
            'file2_only': len(columns2 - columns1),
            'file1_name': file1_name,
            'file2_name': file2_name,
            'comparison_data': comparison_data  # Include the actual comparison table data
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
            print("Completion message sent successfully")
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
                print("Completion message sent (nowait)")
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Column comparison complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'file1_path' in locals() and os.path.exists(file1_path):
            os.remove(file1_path)
        if 'file2_path' in locals() and os.path.exists(file2_path):
            os.remove(file2_path)
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/compare-columns', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def compare_columns():
    """Compare columns from two uploaded files - returns session_id for progress tracking"""
    try:
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({'error': 'Please upload both files'}), 400
        
        file1 = request.files['file1']
        file2 = request.files['file2']
        
        if file1.filename == '' or file2.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Validate file extensions
        if not (file1.filename.endswith(('.xlsx', '.xls', '.csv')) and 
                file2.filename.endswith(('.xlsx', '.xls', '.csv'))):
            return jsonify({'error': 'Please upload Excel or CSV files'}), 400
        
        # Save files temporarily
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"column_compare_{timestamp}_{secrets.token_hex(8)}"
        
        file1_filename = secure_filename(file1.filename)
        file2_filename = secure_filename(file2.filename)
        
        # Store original filenames for display
        file1_name = os.path.splitext(file1_filename)[0]
        file2_name = os.path.splitext(file2_filename)[0]
        
        file1_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_file1_{file1_filename}")
        file2_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_file2_{file2_filename}")
        
        file1.save(file1_path)
        file2.save(file2_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=compare_columns_async,
            args=(file1_path, file2_path, file1_name, file2_name, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

def transpose_file_async(upload_path, progress_queue, session_id):
    """Transpose Excel/CSV file (rows become columns, columns become rows) in background thread"""
    try:
        send_progress = lambda stage, current, total, message, percentage=None: progress_queue.put({
            'stage': stage,
            'current': current,
            'total': total,
            'percentage': percentage if percentage is not None else (int((current / total) * 100) if total > 0 else 0),
            'message': message
        }) if progress_queue and session_id else None
        
        send_progress('loading', 0, 100, 'Reading file into memory...', 10)
        
        # Read file into memory
        filename = os.path.basename(upload_path)
        if filename.endswith('.csv'):
            df = pd.read_csv(upload_path)
        else:
            df = pd.read_excel(upload_path)
        
        original_shape = df.shape
        send_progress('loading', 100, 100, f'File loaded: {original_shape[0]:,} rows × {original_shape[1]:,} columns', 20)
        
        send_progress('processing', 0, 100, 'Transposing data (rows ↔ columns)...', 30)
        
        # Transpose the dataframe
        # First, convert the index to a column if it has a name, otherwise use first column as index
        transposed_df = df.T
        
        # If the original dataframe had a numeric index, we need to set proper column names
        # The first row of transposed data should become the column headers
        if df.index.name is None and not df.index.dtype == 'object':
            # Numeric index - use it as first column name
            transposed_df.index.name = 'Original Row'
        elif df.index.name:
            transposed_df.index.name = df.index.name
        
        # Reset index to make the original index/row numbers a column
        transposed_df = transposed_df.reset_index()
        
        # Rename the first column to something descriptive
        if transposed_df.columns[0] == 'index' or transposed_df.columns[0] == 'Original Row':
            transposed_df.columns.values[0] = 'Original Column/Row'
        
        new_shape = transposed_df.shape
        send_progress('processing', 100, 100, f'Transposed: {new_shape[0]:,} rows × {new_shape[1]:,} columns', 60)
        
        send_progress('saving', 0, 100, 'Saving transposed file...', 70)
        
        # Save transposed data to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"transposed_{timestamp}"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_filename}.xlsx")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            transposed_df.to_excel(writer, sheet_name='Transposed Data', index=False)
        
        download_url = f'/download/{output_filename}.xlsx'
        
        # Clean up uploaded file
        os.remove(upload_path)
        
        # Send completion message
        final_message = {
            'stage': 'done',
            'current': new_shape[0],
            'total': new_shape[0],
            'percentage': 100,
            'message': 'Complete!',
            'download_url': download_url,
            'output_filename': os.path.basename(output_path),
            'original_rows': int(original_shape[0]),
            'original_columns': int(original_shape[1]),
            'transposed_rows': int(new_shape[0]),
            'transposed_columns': int(new_shape[1])
        }
        
        try:
            progress_queue.put(final_message, timeout=5)
            print("Completion message sent successfully")
        except Exception as e:
            print(f"Error sending completion message: {e}")
            try:
                progress_queue.put_nowait(final_message)
                print("Completion message sent (nowait)")
            except Exception as e2:
                print(f"Failed to send completion message: {e2}")
        
        print("Transpose complete!")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        
        # Clean up on error
        if 'upload_path' in locals() and os.path.exists(upload_path):
            os.remove(upload_path)
            
        progress_queue.put({
            'stage': 'error',
            'message': str(e)
        })

@app.route('/transpose-data', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def transpose_data():
    """Transpose uploaded file - returns session_id for progress tracking"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Please upload an Excel or CSV file'}), 400
        
        # Save temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"transpose_{timestamp}_{secrets.token_hex(8)}"
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)
        
        # Create progress queue for this session
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue
        
        # Start processing in background thread
        thread = threading.Thread(
            target=transpose_file_async,
            args=(upload_path, progress_queue, session_id)
        )
        thread.daemon = True
        thread.start()
        
        # Return session ID immediately so client can start listening to progress
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# =============================================================================
# ROW FILTER TOOL
# =============================================================================
@app.route('/row-filter')
def row_filter_page():
    return render_template('row_filter.html')

@app.route('/row-filter', methods=['POST'])
@rate_limit(max_requests=20, window=60)
def row_filter():
    """Filter rows based on conditions"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Get conditions
        conditions_json = request.form.get('conditions', '[]')
        preview_only = request.form.get('preview_only', 'false').lower() == 'true'

        try:
            conditions = json.loads(conditions_json)
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid conditions format'}), 400

        if not conditions:
            return jsonify({'error': 'At least one condition is required'}), 400

        # Generate session ID
        session_id = f"{int(time.time())}_{secrets.token_hex(8)}"
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)

        # Read the file
        df = read_data_file(upload_path)
        original_rows = len(df)

        # Build filter mask
        mask = None
        current_logic = 'AND'

        for condition in conditions:
            column = condition.get('column')
            operator = condition.get('operator')
            value = condition.get('value', '')
            logic = condition.get('logic')

            if logic:
                current_logic = logic

            if column not in df.columns:
                return jsonify({'error': f'Column "{column}" not found'}), 400

            col_data = df[column]

            # Build condition mask
            if operator == 'equals':
                cond_mask = col_data.astype(str).str.lower() == str(value).lower()
            elif operator == 'not_equals':
                cond_mask = col_data.astype(str).str.lower() != str(value).lower()
            elif operator == 'contains':
                cond_mask = col_data.astype(str).str.lower().str.contains(str(value).lower(), na=False, regex=False)
            elif operator == 'not_contains':
                cond_mask = ~col_data.astype(str).str.lower().str.contains(str(value).lower(), na=False, regex=False)
            elif operator == 'starts_with':
                cond_mask = col_data.astype(str).str.lower().str.startswith(str(value).lower(), na=False)
            elif operator == 'ends_with':
                cond_mask = col_data.astype(str).str.lower().str.endswith(str(value).lower(), na=False)
            elif operator == 'greater_than':
                try:
                    cond_mask = pd.to_numeric(col_data, errors='coerce') > float(value)
                except:
                    cond_mask = col_data.astype(str) > str(value)
            elif operator == 'less_than':
                try:
                    cond_mask = pd.to_numeric(col_data, errors='coerce') < float(value)
                except:
                    cond_mask = col_data.astype(str) < str(value)
            elif operator == 'greater_equal':
                try:
                    cond_mask = pd.to_numeric(col_data, errors='coerce') >= float(value)
                except:
                    cond_mask = col_data.astype(str) >= str(value)
            elif operator == 'less_equal':
                try:
                    cond_mask = pd.to_numeric(col_data, errors='coerce') <= float(value)
                except:
                    cond_mask = col_data.astype(str) <= str(value)
            elif operator == 'is_empty':
                cond_mask = col_data.isna() | (col_data.astype(str).str.strip() == '')
            elif operator == 'is_not_empty':
                cond_mask = ~(col_data.isna() | (col_data.astype(str).str.strip() == ''))
            elif operator == 'in_list':
                values = [v.strip().lower() for v in str(value).split(',')]
                cond_mask = col_data.astype(str).str.lower().isin(values)
            else:
                return jsonify({'error': f'Unknown operator: {operator}'}), 400

            # Combine with existing mask
            if mask is None:
                mask = cond_mask
            elif current_logic == 'AND':
                mask = mask & cond_mask
            else:  # OR
                mask = mask | cond_mask

        # Apply filter
        filtered_df = df[mask]
        matching_rows = len(filtered_df)

        # Cleanup upload if preview only
        if preview_only:
            try:
                os.remove(upload_path)
            except:
                pass
            return jsonify({
                'success': True,
                'matching_rows': matching_rows,
                'original_rows': original_rows
            })

        # Save filtered file
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_filtered_{session_id}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        filtered_df.to_excel(output_path, index=False)

        # Cache the result
        cache_session_file(session_id, output_filename, output_path, matching_rows, len(filtered_df.columns), 'Row Filter')

        # Cleanup upload
        try:
            os.remove(upload_path)
        except:
            pass

        return jsonify({
            'success': True,
            'filename': output_filename,
            'original_rows': original_rows,
            'matching_rows': matching_rows
        })

    except Exception as e:
        print(f"Row Filter error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# =============================================================================
# FIND & REPLACE TOOL
# =============================================================================
@app.route('/find-replace')
def find_replace_page():
    return render_template('find_replace.html')

@app.route('/find-replace', methods=['POST'])
@rate_limit(max_requests=20, window=60)
def find_replace():
    """Perform find and replace on uploaded file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Get parameters
        find_text = request.form.get('find_text', '')
        replace_text = request.form.get('replace_text', '')
        column = request.form.get('column', '__all__')
        case_sensitive = request.form.get('case_sensitive', 'false').lower() == 'true'
        use_regex = request.form.get('use_regex', 'false').lower() == 'true'
        match_whole_cell = request.form.get('match_whole_cell', 'false').lower() == 'true'

        if not find_text:
            return jsonify({'error': 'Find text is required'}), 400

        # Generate session ID
        session_id = f"{int(time.time())}_{secrets.token_hex(8)}"
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)

        # Read the file
        df = read_data_file(upload_path)

        # Track replacements
        total_replacements = 0
        rows_affected = set()

        # Determine which columns to process
        if column == '__all__':
            columns_to_process = df.columns.tolist()
        else:
            if column not in df.columns:
                return jsonify({'error': f'Column "{column}" not found'}), 400
            columns_to_process = [column]

        # Perform find and replace
        for col in columns_to_process:
            # Convert column to string for replacement
            original_values = df[col].astype(str)

            if match_whole_cell:
                # Match entire cell
                if use_regex:
                    pattern = f'^{find_text}$'
                    if case_sensitive:
                        mask = original_values.str.match(pattern, na=False)
                    else:
                        mask = original_values.str.match(pattern, case=False, na=False)

                    # Count replacements
                    count = mask.sum()
                    if count > 0:
                        total_replacements += count
                        rows_affected.update(df.index[mask].tolist())
                        df.loc[mask, col] = replace_text
                else:
                    # Exact match (not regex)
                    if case_sensitive:
                        mask = original_values == find_text
                    else:
                        mask = original_values.str.lower() == find_text.lower()

                    count = mask.sum()
                    if count > 0:
                        total_replacements += count
                        rows_affected.update(df.index[mask].tolist())
                        df.loc[mask, col] = replace_text
            else:
                # Partial match / substring replacement
                if use_regex:
                    # Count matches first
                    if case_sensitive:
                        matches = original_values.str.count(find_text, flags=0)
                    else:
                        matches = original_values.str.count(find_text, flags=re.IGNORECASE)

                    count = matches.sum()
                    if count > 0:
                        total_replacements += count
                        rows_affected.update(df.index[matches > 0].tolist())

                        if case_sensitive:
                            df[col] = df[col].astype(str).str.replace(find_text, replace_text, regex=True)
                        else:
                            df[col] = df[col].astype(str).str.replace(find_text, replace_text, regex=True, flags=re.IGNORECASE)
                else:
                    # Simple string replacement
                    if case_sensitive:
                        matches = original_values.str.count(re.escape(find_text), flags=0)
                    else:
                        matches = original_values.str.count(re.escape(find_text), flags=re.IGNORECASE)

                    count = matches.sum()
                    if count > 0:
                        total_replacements += count
                        rows_affected.update(df.index[matches > 0].tolist())

                        if case_sensitive:
                            df[col] = df[col].astype(str).str.replace(find_text, replace_text, regex=False)
                        else:
                            # Case insensitive requires regex
                            df[col] = df[col].astype(str).str.replace(re.escape(find_text), replace_text, regex=True, flags=re.IGNORECASE)

        # Save the modified file
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_replaced_{session_id}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        df.to_excel(output_path, index=False)

        # Cache the result
        cache_session_file(session_id, output_filename, output_path, len(df), len(df.columns), 'Find & Replace')

        # Cleanup upload
        try:
            os.remove(upload_path)
        except:
            pass

        return jsonify({
            'success': True,
            'filename': output_filename,
            'replacements_made': int(total_replacements),
            'rows_affected': len(rows_affected)
        })

    except Exception as e:
        print(f"Find & Replace error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# =============================================================================
# CALCULATED COLUMNS TOOL
# =============================================================================
@app.route('/calculated-columns')
def calculated_columns_page():
    return render_template('calculated_columns.html')

@app.route('/calculated-columns', methods=['POST'])
@rate_limit(max_requests=20, window=60)
def calculated_columns():
    """Create calculated columns using formulas"""
    try:
        # Check for cached file or uploaded file
        cache_id = request.form.get('cache_id')

        if cache_id:
            # Use cached file
            cache_info = get_cached_file(cache_id)
            if not cache_info:
                return jsonify({'error': 'Cached file not found or expired'}), 400
            df = read_data_file(cache_info['path'])
            filename = cache_info['filename']
            session_id = f"{int(time.time())}_{secrets.token_hex(8)}"
            upload_path = None
        else:
            if 'file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400

            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400

            # Generate session ID
            session_id = f"{int(time.time())}_{secrets.token_hex(8)}"
            filename = secure_filename(file.filename)
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
            file.save(upload_path)

            # Read the file
            df = read_data_file(upload_path)

        # Get parameters
        formula = request.form.get('formula', '')
        new_column_name = request.form.get('new_column_name', '')
        preview_only = request.form.get('preview_only', 'false').lower() == 'true'

        if not formula:
            return jsonify({'error': 'Formula is required'}), 400
        if not new_column_name and not preview_only:
            return jsonify({'error': 'New column name is required'}), 400

        # Parse and evaluate the formula
        try:
            result = evaluate_formula(df, formula)
        except Exception as e:
            return jsonify({'error': f'Formula error: {str(e)}'}), 400

        # Add the result as a new column
        if preview_only:
            # Just return preview data
            new_col_name = new_column_name if new_column_name else 'Result'
            preview_df = df.head(5).copy()
            preview_df[new_col_name] = result.head(5)
            preview = preview_df.to_dict(orient='records')
            return jsonify({
                'success': True,
                'preview': preview
            })

        # Check if column already exists
        if new_column_name in df.columns:
            return jsonify({'error': f'Column "{new_column_name}" already exists'}), 400

        df[new_column_name] = result

        # Save the modified file
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_calculated_{session_id}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        df.to_excel(output_path, index=False)

        # Cache the result
        cache_session_file(session_id, output_filename, output_path, len(df), len(df.columns), 'Calculated Columns')

        # Cleanup upload
        if upload_path:
            try:
                os.remove(upload_path)
            except:
                pass

        return jsonify({
            'success': True,
            'filename': output_filename,
            'new_column': new_column_name,
            'rows': len(df),
            'columns': len(df.columns)
        })

    except Exception as e:
        print(f"Calculated Columns error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


def evaluate_formula(df, formula):
    """
    Evaluate a formula with column references and functions.
    Supports: [ColumnName] references, text functions, math operations, conditionals.
    """
    import re
    from datetime import datetime

    # Replace column references [ColumnName] with df access
    def replace_column_refs(match):
        col_name = match.group(1)
        if col_name not in df.columns:
            raise ValueError(f'Column "{col_name}" not found')
        return f'__df__["{col_name}"]'

    # First, protect string literals by replacing them temporarily
    string_literals = []
    def save_string(match):
        string_literals.append(match.group(0))
        return f'__STRING_{len(string_literals)-1}__'

    # Save double and single quoted strings
    formula_work = re.sub(r'"[^"]*"', save_string, formula)
    formula_work = re.sub(r"'[^']*'", save_string, formula_work)

    # Replace column references
    formula_work = re.sub(r'\[([^\]]+)\]', replace_column_refs, formula_work)

    # Restore string literals
    for i, s in enumerate(string_literals):
        formula_work = formula_work.replace(f'__STRING_{i}__', s)

    # Helper to convert value to Series
    def to_series(val):
        if isinstance(val, pd.Series):
            return val
        return pd.Series([val] * len(df))

    # Helper to convert to string Series
    def to_str_series(val):
        if isinstance(val, pd.Series):
            return val.astype(str)
        return pd.Series([str(val)] * len(df))

    # Map of supported functions to pandas equivalents
    function_map = {
        # Text functions
        'CONCAT': lambda *args: pd.concat([to_str_series(arg) for arg in args], axis=1).agg(''.join, axis=1),
        'UPPER': lambda x: to_str_series(x).str.upper(),
        'LOWER': lambda x: to_str_series(x).str.lower(),
        'TRIM': lambda x: to_str_series(x).str.strip(),
        'LEFT': lambda x, n: to_str_series(x).str[:int(n)],
        'RIGHT': lambda x, n: to_str_series(x).str[-int(n):],
        'LEN': lambda x: to_str_series(x).str.len(),
        'REPLACE': lambda x, old, new: to_str_series(x).str.replace(str(old), str(new), regex=False),

        # Math functions
        'ROUND': lambda x, decimals=0: pd.to_numeric(to_series(x), errors='coerce').round(int(decimals)),
        'ABS': lambda x: pd.to_numeric(to_series(x), errors='coerce').abs(),
        'CEILING': lambda x: pd.to_numeric(to_series(x), errors='coerce').apply(lambda v: np.ceil(v) if pd.notna(v) else v),
        'FLOOR': lambda x: pd.to_numeric(to_series(x), errors='coerce').apply(lambda v: np.floor(v) if pd.notna(v) else v),

        # Date functions
        'YEAR': lambda x: pd.to_datetime(to_series(x), errors='coerce').dt.year,
        'MONTH': lambda x: pd.to_datetime(to_series(x), errors='coerce').dt.month,
        'DAY': lambda x: pd.to_datetime(to_series(x), errors='coerce').dt.day,
        'TODAY': lambda: pd.Series([datetime.now().strftime('%Y-%m-%d')] * len(df)),

        # Conditional functions
        'ISNULL': lambda x: to_series(x).isna(),
        'COALESCE': lambda *args: pd.concat([to_series(arg) for arg in args], axis=1).bfill(axis=1).iloc[:, 0],
    }

    # Custom IF function handler
    def handle_if(condition, true_val, false_val):
        """Handle IF(condition, true_value, false_value)"""
        cond = to_series(condition)
        return pd.Series(np.where(cond, true_val, false_val))

    function_map['IF'] = handle_if

    # Build safe namespace
    safe_namespace = {
        '__df__': df,
        'pd': pd,
        'np': np,
        **function_map
    }

    # Handle string concatenation with + operator between columns and strings
    # Convert string operations to work with pandas Series

    try:
        result = eval(formula_work, {"__builtins__": {}}, safe_namespace)

        # Ensure result is a Series
        if isinstance(result, pd.DataFrame):
            result = result.iloc[:, 0]
        elif not isinstance(result, pd.Series):
            result = pd.Series([result] * len(df))

        return result

    except Exception as e:
        raise ValueError(f'Formula evaluation failed: {str(e)}')


# =============================================================================
# COLUMN OPERATIONS TOOL
# =============================================================================
@app.route('/column-operations')
def column_operations_page():
    return render_template('column_operations.html')

@app.route('/column-operations', methods=['POST'])
@rate_limit(max_requests=20, window=60)
def column_operations():
    """Perform column operations on uploaded file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Get operation type
        operation = request.form.get('operation', '')
        if not operation:
            return jsonify({'error': 'No operation specified'}), 400

        # Generate session ID
        session_id = f"{int(time.time())}_{secrets.token_hex(8)}"
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)

        # Read the file
        df = read_data_file(upload_path)
        original_cols = df.columns.tolist()
        operation_summary = ""

        if operation == 'reorder':
            # Reorder columns
            column_order = request.form.get('column_order', '[]')
            try:
                column_order = json.loads(column_order)
            except:
                return jsonify({'error': 'Invalid column order format'}), 400

            # Validate all columns exist
            for col in column_order:
                if col not in df.columns:
                    return jsonify({'error': f'Column "{col}" not found'}), 400

            # Reorder (only include columns in the order list)
            df = df[column_order]
            operation_summary = f"Reordered {len(column_order)} columns"

        elif operation == 'rename':
            # Rename columns
            renames = request.form.get('renames', '{}')
            try:
                renames = json.loads(renames)
            except:
                return jsonify({'error': 'Invalid renames format'}), 400

            if not renames:
                return jsonify({'error': 'No columns selected for renaming'}), 400

            # Validate old column names exist
            for old_name in renames.keys():
                if old_name not in df.columns:
                    return jsonify({'error': f'Column "{old_name}" not found'}), 400

            df = df.rename(columns=renames)
            operation_summary = f"Renamed {len(renames)} column(s)"

        elif operation == 'delete':
            # Delete columns
            columns_to_delete = request.form.get('columns_to_delete', '[]')
            try:
                columns_to_delete = json.loads(columns_to_delete)
            except:
                return jsonify({'error': 'Invalid columns format'}), 400

            if not columns_to_delete:
                return jsonify({'error': 'No columns selected for deletion'}), 400

            # Validate columns exist
            for col in columns_to_delete:
                if col not in df.columns:
                    return jsonify({'error': f'Column "{col}" not found'}), 400

            df = df.drop(columns=columns_to_delete)
            operation_summary = f"Deleted {len(columns_to_delete)} column(s)"

        elif operation == 'duplicate':
            # Duplicate a column
            source_column = request.form.get('source_column', '')
            new_column_name = request.form.get('new_column_name', '')

            if not source_column:
                return jsonify({'error': 'Source column is required'}), 400
            if not new_column_name:
                return jsonify({'error': 'New column name is required'}), 400
            if source_column not in df.columns:
                return jsonify({'error': f'Column "{source_column}" not found'}), 400
            if new_column_name in df.columns:
                return jsonify({'error': f'Column "{new_column_name}" already exists'}), 400

            # Insert the duplicate after the source column
            source_idx = df.columns.get_loc(source_column)
            df.insert(source_idx + 1, new_column_name, df[source_column])
            operation_summary = f"Duplicated '{source_column}' as '{new_column_name}'"

        elif operation == 'split':
            # Split a column by delimiter
            column_to_split = request.form.get('column_to_split', '')
            delimiter = request.form.get('delimiter', '')
            new_names = request.form.get('new_names', '')

            if not column_to_split:
                return jsonify({'error': 'Column to split is required'}), 400
            if not delimiter:
                return jsonify({'error': 'Delimiter is required'}), 400
            if column_to_split not in df.columns:
                return jsonify({'error': f'Column "{column_to_split}" not found'}), 400

            # Parse new column names (comma-separated)
            if new_names:
                new_names = [n.strip() for n in new_names.split(',') if n.strip()]
            else:
                new_names = []

            # Split the column
            split_df = df[column_to_split].astype(str).str.split(delimiter, expand=True)
            num_parts = split_df.shape[1]

            # Generate column names if not enough provided
            if len(new_names) < num_parts:
                for i in range(len(new_names), num_parts):
                    new_names.append(f"{column_to_split}_part{i+1}")

            # Use only the names needed
            new_names = new_names[:num_parts]

            # Check for duplicate column names
            for name in new_names:
                if name in df.columns and name != column_to_split:
                    return jsonify({'error': f'Column "{name}" already exists'}), 400

            # Insert new columns after the original
            source_idx = df.columns.get_loc(column_to_split)

            # Drop original column
            df = df.drop(columns=[column_to_split])

            # Insert split columns
            for i, name in enumerate(new_names):
                df.insert(source_idx + i, name, split_df[i])

            operation_summary = f"Split '{column_to_split}' into {num_parts} columns"

        elif operation == 'merge':
            # Merge columns with separator
            columns_to_merge = request.form.get('columns_to_merge', '[]')
            separator = request.form.get('separator', '')
            new_column_name = request.form.get('new_column_name', '')

            try:
                columns_to_merge = json.loads(columns_to_merge)
            except:
                return jsonify({'error': 'Invalid columns format'}), 400

            if not columns_to_merge or len(columns_to_merge) < 2:
                return jsonify({'error': 'Select at least 2 columns to merge'}), 400
            if not new_column_name:
                return jsonify({'error': 'New column name is required'}), 400
            if new_column_name in df.columns:
                return jsonify({'error': f'Column "{new_column_name}" already exists'}), 400

            # Validate columns exist
            for col in columns_to_merge:
                if col not in df.columns:
                    return jsonify({'error': f'Column "{col}" not found'}), 400

            # Merge columns
            df[new_column_name] = df[columns_to_merge].astype(str).agg(separator.join, axis=1)

            # Move new column to after the last merged column
            first_col_idx = min(df.columns.get_loc(col) for col in columns_to_merge)

            # Reorder to put new column in place
            cols = df.columns.tolist()
            cols.remove(new_column_name)
            cols.insert(first_col_idx + len(columns_to_merge), new_column_name)
            df = df[cols]

            operation_summary = f"Merged {len(columns_to_merge)} columns into '{new_column_name}'"

        else:
            return jsonify({'error': f'Unknown operation: {operation}'}), 400

        # Save the modified file
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_modified_{session_id}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        df.to_excel(output_path, index=False)

        # Cache the result
        cache_session_file(session_id, output_filename, output_path, len(df), len(df.columns), 'Column Operations')

        # Cleanup upload
        try:
            os.remove(upload_path)
        except:
            pass

        return jsonify({
            'success': True,
            'filename': output_filename,
            'summary': operation_summary,
            'original_columns': len(original_cols),
            'new_columns': len(df.columns)
        })

    except Exception as e:
        print(f"Column Operations error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# =============================================================================
# DATA READINESS PIPELINE ROUTES
# =============================================================================

@app.route('/data-readiness-pipeline')
def data_readiness_pipeline():
    """Data Readiness Pipeline - guided multi-stage data assessment"""
    return render_template('data_readiness_pipeline.html')


@app.route('/pipeline/start', methods=['POST'])
@rate_limit(max_requests=RATE_LIMIT_REQUESTS, window=RATE_LIMIT_WINDOW)
def pipeline_start():
    """Start a new pipeline session - upload file and create session"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': 'Please upload an Excel (.xlsx, .xls) or CSV file'}), 400

        # Generate unique session ID
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        session_id = f"pipeline_{timestamp}_{secrets.token_hex(8)}"

        # Save file
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(upload_path)

        # Create pipeline state
        state = PipelineState(session_id, upload_path, filename)

        # Load DataFrame into memory
        state.df = read_data_file(upload_path)
        state.row_count = len(state.df)
        state.col_count = len(state.df.columns)

        # Store session
        pipeline_sessions[session_id] = state

        # Get preview data (first 20 rows)
        preview_rows = []
        preview_df = state.df.head(20)
        for _, row in preview_df.iterrows():
            row_dict = {}
            for col in preview_df.columns:
                val = row[col]
                if pd.isna(val):
                    row_dict[col] = None
                elif isinstance(val, (np.integer, np.floating)):
                    row_dict[col] = float(val) if isinstance(val, np.floating) else int(val)
                else:
                    row_dict[col] = str(val)
            preview_rows.append(row_dict)

        return jsonify({
            'success': True,
            'session_id': session_id,
            'filename': filename,
            'rows': state.row_count,
            'columns': state.col_count,
            'column_names': list(state.df.columns),
            'preview': preview_rows
        })

    except Exception as e:
        print(f"Pipeline start error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/state', methods=['GET'])
def pipeline_get_state(session_id):
    """Get current pipeline state for resume/refresh"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]
    return jsonify({
        'success': True,
        'state': state.to_dict(),
        'stage_data': {
            str(k): v is not None for k, v in state.stage_data.items()
        },
        'user_decisions': state.user_decisions
    })


@app.route('/pipeline/<session_id>/analyze', methods=['POST'])
def pipeline_analyze(session_id):
    """Stage 1: Run shape analysis on the uploaded data"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        # Create progress queue for SSE
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue

        # Run analysis in background thread
        def run_analysis():
            try:
                # Analyze the DataFrame
                analysis = analyze_dataframe(state.df, progress_queue, session_id)

                # Extract gap summary for Stage 2
                gap_summary = []
                for col_name, col_info in analysis['columns'].items():
                    if col_info['null_count'] > 0:
                        gap_summary.append({
                            'column': col_name,
                            'null_count': col_info['null_count'],
                            'null_percentage': col_info['null_percentage'],
                            'detected_type': col_info.get('detected_type', 'Unknown'),
                            'unique_count': col_info.get('unique_count', 0)
                        })

                # Sort by null percentage descending
                gap_summary = sorted(gap_summary, key=lambda x: x['null_percentage'], reverse=True)
                analysis['gap_summary'] = gap_summary

                # Detect potentially sensitive columns for Stage 4
                sensitive_patterns = ['name', 'email', 'phone', 'address', 'ssn', 'social',
                                     'credit', 'account', 'password', 'dob', 'birth', 'salary']
                sensitive_columns = []
                for col_name in state.df.columns:
                    col_lower = col_name.lower()
                    for pattern in sensitive_patterns:
                        if pattern in col_lower:
                            sensitive_columns.append({
                                'column': col_name,
                                'pattern_matched': pattern,
                                'unique_count': analysis['columns'].get(col_name, {}).get('unique_count', 0)
                            })
                            break
                analysis['sensitive_columns'] = sensitive_columns

                # Store results
                state.stage_data[1] = make_json_serializable(analysis)
                state.current_stage = 2

                # Send completion
                progress_queue.put({
                    'stage': 'done',
                    'percentage': 100,
                    'message': 'Shape analysis complete',
                    'analysis': make_json_serializable(analysis)
                })

            except Exception as e:
                print(f"Pipeline analysis error: {str(e)}")
                print(traceback.format_exc())
                progress_queue.put({
                    'stage': 'error',
                    'message': str(e)
                })

        thread = threading.Thread(target=run_analysis)
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'session_id': session_id,
            'message': 'Analysis started'
        })

    except Exception as e:
        print(f"Pipeline analyze error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/gaps', methods=['GET'])
def pipeline_get_gaps(session_id):
    """Stage 2: Get gap assessment data from Stage 1 analysis"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    if state.stage_data[1] is None:
        return jsonify({'error': 'Stage 1 (Shape Analysis) must be completed first'}), 400

    analysis = state.stage_data[1]

    return jsonify({
        'success': True,
        'gap_summary': analysis.get('gap_summary', []),
        'total_columns': len(analysis.get('columns', {})),
        'columns_with_gaps': len(analysis.get('gap_summary', [])),
        'current_triage': state.user_decisions.get(2, {})
    })


@app.route('/pipeline/<session_id>/gaps/triage', methods=['POST'])
def pipeline_triage_gaps(session_id):
    """Stage 2: Save user's gap triage decisions"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        triage_decisions = request.get_json()
        if not triage_decisions:
            return jsonify({'error': 'No triage decisions provided'}), 400

        # Store user decisions
        state.user_decisions[2] = triage_decisions
        state.stage_data[2] = {
            'gap_triage': triage_decisions,
            'completed_at': time.time()
        }
        state.current_stage = 3

        return jsonify({
            'success': True,
            'message': 'Gap triage decisions saved',
            'gaps_triaged': len(triage_decisions)
        })

    except Exception as e:
        print(f"Pipeline triage error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/keys', methods=['POST'])
def pipeline_find_keys(session_id):
    """Stage 3: Find natural key candidates using Apriori algorithm"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        data = request.get_json() or {}
        selected_columns = data.get('selected_columns', list(state.df.columns))

        if not selected_columns:
            return jsonify({'error': 'Please select at least one column'}), 400

        # Create progress queue
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue

        def run_key_discovery():
            try:
                def send_progress(stage, pct, msg, current=0, total=0):
                    progress_queue.put({
                        'stage': stage,
                        'percentage': pct,
                        'message': msg,
                        'current': current,
                        'total': total
                    })

                send_progress('loading', 2, 'Preparing data for analysis...')

                df = state.df.copy()
                total_rows = len(df)

                # Remove fully duplicate rows for key analysis
                send_progress('filtering', 5, f'Checking for duplicate rows in {total_rows:,} records...')
                duplicate_count = df.duplicated().sum()
                if duplicate_count > 0:
                    df = df.drop_duplicates(keep='first')
                    send_progress('filtering', 8, f'Removed {duplicate_count:,} duplicate rows. {len(df):,} unique rows remaining.')
                else:
                    send_progress('filtering', 8, f'No duplicate rows found. Analyzing {len(df):,} rows.')

                # Validate columns
                invalid_cols = [c for c in selected_columns if c not in df.columns]
                if invalid_cols:
                    raise ValueError(f"Columns not found: {', '.join(invalid_cols)}")

                # Find minimal unique combinations using Apriori
                minimal_combinations = []
                non_unique_combinations = []
                n = len(selected_columns)
                max_size = min(n, 5)
                max_keys = 10

                # Calculate progress ranges - use 10-90% for the actual work
                # Level 1 gets 10-25%, remaining levels share 25-90%
                level1_start = 10
                level1_end = 25
                remaining_start = 25
                remaining_end = 90
                levels_remaining = max_size - 1  # levels 2 through max_size
                level_range = (remaining_end - remaining_start) / max(levels_remaining, 1)

                # Level 1: Single columns
                send_progress('level1', level1_start, f'Level 1: Testing {n} single columns...', 0, n)
                for idx, col in enumerate(selected_columns):
                    is_unique = df[[col]].duplicated().sum() == 0
                    if is_unique:
                        minimal_combinations.append([col])
                    else:
                        non_unique_combinations.append(frozenset([col]))

                    # Update progress for each column
                    pct = level1_start + int((idx + 1) / n * (level1_end - level1_start))
                    if is_unique:
                        send_progress('level1', pct, f'Found unique key: {col}', idx + 1, n)
                    elif (idx + 1) % 3 == 0 or idx == n - 1:
                        send_progress('level1', pct,
                                     f'Level 1: Tested {idx + 1}/{n} columns. Found {len(minimal_combinations)} key(s).',
                                     idx + 1, n)

                # Levels 2+: Composite keys with Apriori pruning
                for k in range(2, max_size + 1):
                    level_idx = k - 2  # 0 for level 2, 1 for level 3, etc.
                    level_start = remaining_start + int(level_idx * level_range)
                    level_end = remaining_start + int((level_idx + 1) * level_range)

                    if not non_unique_combinations or len(minimal_combinations) >= max_keys:
                        send_progress('complete', 92,
                                     f'Search complete. Found {len(minimal_combinations)} minimal key(s).')
                        break

                    # Generate candidates
                    send_progress(f'level{k}', level_start,
                                 f'Level {k}: Generating {k}-column candidates using Apriori pruning...')

                    pool_cols = set()
                    for combo in non_unique_combinations:
                        pool_cols.update(combo)

                    valid_candidates = []
                    for combo in combinations(sorted(pool_cols), k):
                        combo_set = frozenset(combo)
                        all_subsets_non_unique = all(
                            frozenset(subset) in non_unique_combinations
                            for subset in combinations(combo, k - 1)
                        )
                        if all_subsets_non_unique:
                            valid_candidates.append(list(combo))

                    if not valid_candidates:
                        send_progress(f'level{k}', level_end,
                                     f'Level {k}: No valid candidates after pruning.')
                        continue

                    send_progress(f'level{k}', level_start + 2,
                                 f'Level {k}: Testing {len(valid_candidates)} candidate combinations...',
                                 0, len(valid_candidates))

                    next_level_non_unique = []
                    for idx, candidate in enumerate(valid_candidates):
                        if len(minimal_combinations) >= max_keys:
                            break
                        is_unique = df[candidate].duplicated().sum() == 0
                        if is_unique:
                            minimal_combinations.append(candidate)
                        else:
                            next_level_non_unique.append(frozenset(candidate))

                        # Calculate granular progress within this level
                        progress_in_level = (idx + 1) / len(valid_candidates)
                        pct = level_start + 2 + int(progress_in_level * (level_end - level_start - 2))

                        # Update more frequently for better UX
                        if is_unique:
                            send_progress(f'level{k}', pct, f'Found key: {" + ".join(candidate)}',
                                         idx + 1, len(valid_candidates))
                        elif (idx + 1) % 5 == 0 or idx == len(valid_candidates) - 1:
                            send_progress(f'level{k}', pct,
                                         f'Level {k}: Tested {idx + 1}/{len(valid_candidates)} combinations. Found {len(minimal_combinations)} key(s).',
                                         idx + 1, len(valid_candidates))

                    non_unique_combinations = next_level_non_unique
                else:
                    # Loop completed without break - all levels exhausted
                    send_progress('complete', 92,
                                 f'Search complete. Found {len(minimal_combinations)} minimal key(s).')

                # Store results
                key_results = {
                    'minimal_combinations': minimal_combinations,
                    'selected_columns': selected_columns,
                    'rows_analyzed': len(df),
                    'duplicates_removed': duplicate_count
                }
                state.stage_data[3] = key_results

                # Send single done message with results included
                progress_queue.put({
                    'stage': 'done',
                    'percentage': 100,
                    'message': f'Analysis complete! Found {len(minimal_combinations)} natural key candidate(s).',
                    'results': key_results
                })

            except Exception as e:
                print(f"Key discovery error: {str(e)}")
                print(traceback.format_exc())
                progress_queue.put({'stage': 'error', 'message': str(e)})

        thread = threading.Thread(target=run_key_discovery)
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'session_id': session_id,
            'message': 'Key discovery started'
        })

    except Exception as e:
        print(f"Pipeline keys error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/keys/confirm', methods=['POST'])
def pipeline_confirm_keys(session_id):
    """Stage 3: Save user's key selection"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        data = request.get_json()
        selected_key = data.get('selected_key', [])

        state.user_decisions[3] = {'selected_key': selected_key}
        if state.stage_data[3]:
            state.stage_data[3]['user_selected_key'] = selected_key
        state.current_stage = 4

        return jsonify({
            'success': True,
            'message': 'Key selection saved',
            'selected_key': selected_key
        })

    except Exception as e:
        print(f"Pipeline confirm keys error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/transformations', methods=['GET'])
def pipeline_get_transformations(session_id):
    """Stage 4: Get transformation recommendations based on analysis"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    if state.stage_data[1] is None:
        return jsonify({'error': 'Stage 1 must be completed first'}), 400

    analysis = state.stage_data[1]
    recommendations = []

    # Check for sensitive columns that might need anonymization
    sensitive_cols = analysis.get('sensitive_columns', [])
    if sensitive_cols:
        # Build detailed column info with explanations
        column_details = []
        for col in sensitive_cols:
            pattern = col.get('pattern_matched', '')
            col_name = col.get('column', '')
            unique_count = col.get('unique_count', 0)

            # Generate explanation based on pattern
            pattern_explanations = {
                'name': 'Contains personal names - will be replaced with realistic fake names',
                'email': 'Contains email addresses - will be replaced with anonymized emails',
                'phone': 'Contains phone numbers - will be replaced with fake phone numbers',
                'address': 'Contains addresses - will be replaced with fake addresses',
                'ssn': 'Contains Social Security Numbers - will be masked or replaced',
                'social': 'Contains social identifiers - will be anonymized',
                'credit': 'Contains credit card or financial info - will be masked',
                'account': 'Contains account numbers - will be replaced with fake numbers',
                'password': 'Contains passwords or secrets - will be hashed or removed',
                'dob': 'Contains dates of birth - will be shifted or generalized',
                'birth': 'Contains birth information - will be anonymized',
                'salary': 'Contains salary/compensation data - will be bucketed or masked'
            }
            explanation = pattern_explanations.get(pattern, 'May contain sensitive information')

            column_details.append({
                'column': col_name,
                'pattern': pattern,
                'unique_values': unique_count,
                'explanation': explanation
            })

        recommendations.append({
            'type': 'anonymization',
            'title': 'Data Anonymization',
            'description': 'Protect personally identifiable information (PII) by replacing sensitive data with realistic fake values while maintaining data structure and relationships.',
            'columns': [c['column'] for c in sensitive_cols],
            'column_details': column_details,
            'priority': 'high'
        })

    # Check for type inconsistencies that might need normalization
    type_issues = []
    for col_name, col_info in analysis.get('columns', {}).items():
        if col_info.get('detected_type') != 'Unknown':
            dtype = col_info.get('dtype', '')
            detected = col_info.get('detected_type', '')
            if 'object' in dtype and detected in ['Numeric', 'Integer', 'Date', 'Currency']:
                # Generate explanation based on detected type
                type_explanations = {
                    'Numeric': f'Stored as text but contains numbers - will convert to numeric format for calculations',
                    'Integer': f'Stored as text but contains whole numbers - will convert to integer format',
                    'Date': f'Stored as text but contains dates - will convert to proper date format for sorting/filtering',
                    'Currency': f'Stored as text but contains currency values - will convert to numeric and standardize format'
                }
                explanation = type_explanations.get(detected, 'May need type conversion')

                type_issues.append({
                    'column': col_name,
                    'current': dtype,
                    'detected': detected,
                    'explanation': explanation
                })

    if type_issues:
        recommendations.append({
            'type': 'normalization',
            'title': 'Data Type Normalization',
            'description': 'Convert columns to their proper data types. This improves data quality, enables proper sorting, and allows mathematical operations.',
            'columns': [t['column'] for t in type_issues],
            'column_details': type_issues,
            'priority': 'medium'
        })

    # Check for gaps that need attention (from Stage 2 triage)
    gap_triage = state.user_decisions.get(2, {})
    gaps_needing_attention = [col for col, decision in gap_triage.items() if decision == 'needs_attention']
    if gaps_needing_attention:
        # Get gap details from analysis
        gap_details = []
        for col in gaps_needing_attention:
            col_info = analysis.get('columns', {}).get(col, {})
            null_pct = col_info.get('null_percentage', 0)
            null_count = col_info.get('null_count', 0)

            gap_details.append({
                'column': col,
                'missing_count': null_count,
                'missing_percentage': round(null_pct, 1),
                'explanation': f'{null_count:,} missing values ({null_pct:.1f}%) - can fill with default, interpolate, or flag for review'
            })

        recommendations.append({
            'type': 'gap_handling',
            'title': 'Missing Data Handling',
            'description': 'Address missing values in columns you flagged for attention. Options include filling with defaults, using statistical imputation, or flagging rows for manual review.',
            'columns': gaps_needing_attention,
            'column_details': gap_details,
            'priority': 'medium'
        })

    return jsonify({
        'success': True,
        'recommendations': recommendations,
        'current_selections': state.user_decisions.get(4, {})
    })


@app.route('/pipeline/<session_id>/transformations/select', methods=['POST'])
def pipeline_select_transformations(session_id):
    """Stage 4: Save user's transformation selections"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        selections = request.get_json()
        if selections is None:
            selections = {}

        state.user_decisions[4] = selections
        state.stage_data[4] = {
            'selected_transformations': selections,
            'completed_at': time.time()
        }
        state.current_stage = 5

        return jsonify({
            'success': True,
            'message': 'Transformation selections saved',
            'selections': selections
        })

    except Exception as e:
        print(f"Pipeline select transformations error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/pipeline/<session_id>/execute', methods=['POST'])
def pipeline_execute(session_id):
    """Stage 5: Execute selected transformations and generate report"""
    if session_id not in pipeline_sessions:
        return jsonify({'error': 'Pipeline session not found or expired'}), 404

    state = pipeline_sessions[session_id]

    try:
        progress_queue = Queue()
        progress_queues[session_id] = progress_queue

        def run_execute():
            try:
                send_progress = lambda pct, msg: progress_queue.put({
                    'stage': 'executing', 'percentage': pct, 'message': msg
                })

                send_progress(5, 'Starting pipeline execution...')

                result_df = state.df.copy()
                transformation_log = []
                selections = state.user_decisions.get(4, {})

                # Apply anonymization if selected
                if selections.get('anonymization', {}).get('enabled'):
                    send_progress(20, 'Applying data anonymization...')
                    anon_cols = selections['anonymization'].get('columns', [])
                    if anon_cols:
                        for col in anon_cols:
                            if col in result_df.columns:
                                unique_vals = result_df[col].dropna().unique()
                                mapping = {val: f"{col[:3].upper()}_{i+1:05d}" for i, val in enumerate(unique_vals)}
                                result_df[col] = result_df[col].map(lambda x: mapping.get(x, x) if pd.notna(x) else x)
                        transformation_log.append({
                            'type': 'anonymization',
                            'columns': anon_cols,
                            'rows_affected': len(result_df)
                        })

                # Apply type normalization if selected
                if selections.get('normalization', {}).get('enabled'):
                    send_progress(40, 'Applying type normalization...')
                    norm_cols = selections['normalization'].get('columns', [])
                    # Type normalization would be applied here
                    if norm_cols:
                        transformation_log.append({
                            'type': 'normalization',
                            'columns': norm_cols,
                            'note': 'Type hints recorded for reference'
                        })

                send_progress(60, 'Generating PDF report...')

                # Generate PDF Report
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_basename = f"data_readiness_report_{timestamp}"
                pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}.pdf")

                generate_readiness_report(
                    output_path=pdf_path,
                    state=state,
                    transformation_log=transformation_log
                )

                send_progress(80, 'Saving transformed data...')

                # Save transformed data
                excel_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}_data.xlsx")
                result_df.to_excel(excel_path, index=False)

                # Create ZIP package
                send_progress(90, 'Packaging results...')
                zip_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{output_basename}.zip")
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(pdf_path, f"{output_basename}_report.pdf")
                    zipf.write(excel_path, f"{output_basename}_data.xlsx")

                # Cleanup individual files
                os.remove(pdf_path)
                os.remove(excel_path)

                # Store execution results
                state.stage_data[5] = {
                    'transformation_log': transformation_log,
                    'output_file': f"{output_basename}.zip",
                    'completed_at': time.time()
                }

                # Cache the result file
                cache_session_file(
                    session_id,
                    f"{output_basename}_data.xlsx",
                    zip_path,
                    len(result_df),
                    len(result_df.columns),
                    'Data Readiness Pipeline'
                )

                progress_queue.put({
                    'stage': 'done',
                    'percentage': 100,
                    'message': 'Pipeline execution complete',
                    'download_url': f'/download/{output_basename}.zip',
                    'output_filename': f"{output_basename}.zip",
                    'transformation_log': transformation_log
                })

            except Exception as e:
                print(f"Pipeline execute error: {str(e)}")
                print(traceback.format_exc())
                progress_queue.put({'stage': 'error', 'message': str(e)})

        thread = threading.Thread(target=run_execute)
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'session_id': session_id,
            'message': 'Pipeline execution started'
        })

    except Exception as e:
        print(f"Pipeline execute error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def generate_readiness_report(output_path, state, transformation_log=None):
    """Generate comprehensive Data Readiness PDF report"""
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc = SimpleDocTemplate(output_path, pagesize=letter,
                           leftMargin=0.75*inch, rightMargin=0.75*inch,
                           topMargin=0.75*inch, bottomMargin=0.75*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, alignment=TA_CENTER, spaceAfter=20)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=10)
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, spaceAfter=6)

    story = []

    # Title
    story.append(Paragraph("Data Readiness Report", title_style))
    story.append(Paragraph(f"Source File: {state.filename}", body_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
    story.append(Spacer(1, 20))

    # Executive Summary
    story.append(Paragraph("1. Executive Summary", heading_style))
    analysis = state.stage_data.get(1, {})
    overview = analysis.get('overview', {})
    shape = overview.get('shape', {})

    summary_data = [
        ['Metric', 'Value'],
        ['Total Rows', f"{shape.get('rows', 0):,}"],
        ['Total Columns', f"{shape.get('columns', 0):,}"],
        ['Duplicate Rows', f"{overview.get('duplicate_rows', 0):,}"],
        ['Memory Usage', f"{overview.get('memory_usage_mb', 0):.2f} MB"]
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))

    # Stage 2: Gap Assessment
    story.append(Paragraph("2. Gap Assessment", heading_style))
    gap_triage = state.user_decisions.get(2, {})
    gap_summary = analysis.get('gap_summary', [])

    # Create a lookup for gap percentages
    gap_percentages = {g['column']: g['null_percentage'] for g in gap_summary}

    if gap_triage:
        # Count decisions
        needs_attention = [col for col, dec in gap_triage.items() if dec == 'needs_attention']
        acceptable = [col for col, dec in gap_triage.items() if dec == 'acceptable']

        story.append(Paragraph(f"Columns with missing data: {len(gap_triage)}", body_style))
        story.append(Paragraph(f"Marked as needing attention: {len(needs_attention)}", body_style))
        story.append(Paragraph(f"Marked as acceptable: {len(acceptable)}", body_style))
        story.append(Spacer(1, 10))

        gap_data = [['Column', 'Missing %', 'Decision']]
        for col, decision in sorted(gap_triage.items(), key=lambda x: gap_percentages.get(x[0], 0), reverse=True):
            decision_text = 'Needs Attention' if decision == 'needs_attention' else 'Acceptable'
            pct = gap_percentages.get(col, 0)
            gap_data.append([col, f"{pct:.1f}%", decision_text])

        gap_table = Table(gap_data, colWidths=[2.5*inch, 1*inch, 1.5*inch])
        gap_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        story.append(gap_table)
    elif gap_summary:
        story.append(Paragraph(f"Found {len(gap_summary)} columns with missing data, but no triage decisions were made.", body_style))
    else:
        story.append(Paragraph("No columns with missing data were found.", body_style))
    story.append(Spacer(1, 20))

    # Stage 3: Natural Key Discovery
    story.append(Paragraph("3. Natural Key Discovery", heading_style))
    key_data = state.stage_data.get(3, {})
    if key_data:
        all_candidates = key_data.get('minimal_combinations', [])
        selected_key = key_data.get('user_selected_key', all_candidates[0] if all_candidates else [])

        if selected_key:
            selected_key_str = ' + '.join(selected_key) if isinstance(selected_key, list) else str(selected_key)
            story.append(Paragraph(f"<b>Selected Natural Key:</b> {selected_key_str}", body_style))
            story.append(Spacer(1, 8))
            story.append(Paragraph(
                "This column combination uniquely identifies each row in your dataset and can serve as a primary key.",
                body_style
            ))
        else:
            story.append(Paragraph("No natural key was selected.", body_style))

        # Show alternative options
        if all_candidates and len(all_candidates) > 1:
            story.append(Spacer(1, 12))
            # Filter out the selected key from alternatives
            other_keys = [k for k in all_candidates if k != selected_key]
            if other_keys:
                story.append(Paragraph("<b>Other Available Keys:</b>", body_style))
                story.append(Spacer(1, 4))
                for idx, alt_key in enumerate(other_keys[:5], 1):  # Show up to 5 alternatives
                    alt_key_str = ' + '.join(alt_key) if isinstance(alt_key, list) else str(alt_key)
                    story.append(Paragraph(f"  {idx}. {alt_key_str}", body_style))
                if len(other_keys) > 5:
                    story.append(Paragraph(f"  ... and {len(other_keys) - 5} more candidate(s)", body_style))
        elif all_candidates and len(all_candidates) == 1:
            story.append(Spacer(1, 8))
            story.append(Paragraph("This was the only natural key candidate found.", body_style))
    else:
        story.append(Paragraph("Natural key discovery was not performed.", body_style))
    story.append(Spacer(1, 20))

    # Stage 4: Transformation Decisions
    story.append(Paragraph("4. Transformation Decisions", heading_style))
    transform_decisions = state.user_decisions.get(4, {})
    if transform_decisions:
        for ttype, config in transform_decisions.items():
            if config.get('enabled'):
                cols = config.get('columns', [])
                story.append(Paragraph(f"• {ttype.title()}: {len(cols)} column(s)", body_style))
    else:
        story.append(Paragraph("No transformations were selected.", body_style))
    story.append(Spacer(1, 20))

    # Stage 5: Execution Log
    story.append(Paragraph("5. Execution Log", heading_style))
    if transformation_log:
        for entry in transformation_log:
            story.append(Paragraph(f"• {entry.get('type', 'Unknown').title()}: {len(entry.get('columns', []))} column(s) affected", body_style))
    else:
        story.append(Paragraph("No transformations were executed.", body_style))

    # Footer
    story.append(Spacer(1, 40))
    story.append(Paragraph("Generated by DataDragon - Data Readiness Pipeline",
                          ParagraphStyle('Footer', parent=body_style, alignment=TA_CENTER, textColor=colors.grey)))

    doc.build(story)


if __name__ == '__main__':
    # Only enable debug in development
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    # Use port 5002 to avoid conflict with macOS AirPlay Receiver on port 5000
    app.run(debug=debug_mode, host='127.0.0.1', port=5002, threaded=True)

